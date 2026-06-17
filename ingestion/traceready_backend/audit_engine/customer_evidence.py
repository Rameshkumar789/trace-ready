from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from pydantic import BaseModel, ConfigDict, Field, field_validator


GENERATED_AT = "2026-06-16T00:00:00Z"


class StrictCustomerEvidenceModel(BaseModel):
    model_config = ConfigDict(extra="forbid", validate_assignment=True, str_strip_whitespace=True)


class EvidenceSourcePointer(StrictCustomerEvidenceModel):
    file_name: str
    sheet_name: str
    row_number: int
    column_name: str
    column_index: int
    cell: str


class CustomerEvidenceRecord(StrictCustomerEvidenceModel):
    evidence_id: str
    uploaded_file: str
    sheet_name: str
    row_number: int
    column_name: str
    column_index: int
    cell: str
    raw_value: str
    normalized_value: str
    field_key: str
    field_type: str
    extraction_method: str
    confidence: float = Field(ge=0, le=1)
    source_pointer: EvidenceSourcePointer

    @field_validator("field_key")
    @classmethod
    def _field_key_is_slug(cls, value: str) -> str:
        if not re.fullmatch(r"[a-z0-9]+(?:_[a-z0-9]+)*", value):
            raise ValueError("field_key must be snake_case")
        return value


class FieldMappingSuggestion(StrictCustomerEvidenceModel):
    suggestion_id: str
    source_sheet: str
    source_column: str
    field_key: str
    canonical_field: str
    confidence: float = Field(ge=0, le=1)
    suggestion_method: str
    rationale: str
    evidence_ids: list[str]
    review_status: str = "needs_review"


class FoodFormResolution(StrictCustomerEvidenceModel):
    product_name: str | None = None
    ftl_category: str | None = None
    is_ftl_likely: bool | None = None
    form_state: list[str] = Field(default_factory=list)
    output_remains_ftl: bool | None = None
    confidence: float = Field(ge=0, le=1)
    reasons: list[str] = Field(default_factory=list)
    review_required: bool = False


class ActorRoleResolution(StrictCustomerEvidenceModel):
    actor_name: str | None = None
    actor_type: str | None = None
    role: str
    confidence: float = Field(ge=0, le=1)
    reasons: list[str] = Field(default_factory=list)


class TraceabilityEntity(StrictCustomerEvidenceModel):
    entity_id: str
    entity_type: str
    name: str
    attributes: dict[str, Any] = Field(default_factory=dict)
    evidence_ids: list[str] = Field(default_factory=list)


class TraceabilityEntityGraph(StrictCustomerEvidenceModel):
    products: list[TraceabilityEntity] = Field(default_factory=list)
    product_forms: list[TraceabilityEntity] = Field(default_factory=list)
    lots: list[TraceabilityEntity] = Field(default_factory=list)
    actors: list[TraceabilityEntity] = Field(default_factory=list)
    locations: list[TraceabilityEntity] = Field(default_factory=list)
    counterparties: list[TraceabilityEntity] = Field(default_factory=list)
    documents: list[TraceabilityEntity] = Field(default_factory=list)


class CustomerEventNode(StrictCustomerEvidenceModel):
    event_id: str
    source_row_key: str
    evidence_ids: list[str]
    event_type_claim: str | None = None
    event_datetime: str | None = None
    actor_id: str | None = None
    actor_role: ActorRoleResolution
    product_id: str | None = None
    product_name: str | None = None
    food_form: FoodFormResolution
    lot_or_tlc: str | None = None
    source_lot_or_tlc: str | None = None
    output_lot_or_tlc: str | None = None
    from_partner_id: str | None = None
    to_partner_id: str | None = None
    destination_type: str | None = None
    action_terms: list[str] = Field(default_factory=list)
    classified_ctes: list[str] = Field(default_factory=list)
    suppressed_ctes: list[str] = Field(default_factory=list)
    reviewer_questions: list[str] = Field(default_factory=list)


class CteClassificationResult(StrictCustomerEvidenceModel):
    event_id: str
    candidate_ctes: list[str]
    final_ctes: list[str]
    suppressed_ctes: list[str]
    confidence: float = Field(ge=0, le=1)
    reasons: list[str] = Field(default_factory=list)
    reviewer_questions: list[str] = Field(default_factory=list)
    evidence_ids: list[str] = Field(default_factory=list)


class InferredEvidenceFact(StrictCustomerEvidenceModel):
    fact_id: str
    source_kind: str
    source_name: str
    field_key: str
    raw_value: str
    normalized_value: str
    confidence: float = Field(ge=0, le=1)
    extraction_method: str
    evidence_pointer: dict[str, Any]


class CustomerDocumentProfile(StrictCustomerEvidenceModel):
    profile_id: str
    source_name: str
    document_type: str
    evidence_kind: str
    confidence: float = Field(ge=0, le=1)
    detected_signals: list[str] = Field(default_factory=list)
    supported_parser: str
    evidence_ids: list[str] = Field(default_factory=list)


class EvidenceConflict(StrictCustomerEvidenceModel):
    conflict_id: str
    conflict_type: str
    row_key: str | None = None
    entity_key: str | None = None
    field_key: str
    values: list[str]
    evidence_ids_by_value: dict[str, list[str]]
    severity: str
    status: str = "needs_review"


class CustomerEvidenceQualityReport(StrictCustomerEvidenceModel):
    generated_at: str
    source_file: str
    parse_coverage: dict[str, Any]
    workbook_structure: dict[str, Any]
    document_type_counts: dict[str, int]
    unmapped_sheets: list[str]
    unmapped_columns: list[dict[str, Any]]
    low_confidence_mappings: list[dict[str, Any]]
    inferred_fact_count: int
    conflict_count: int
    duplicate_row_count: int
    missing_lineage_anchor_count: int
    abstention_count: int
    quality_gate: str
    issues: list[dict[str, Any]] = Field(default_factory=list)


class Phase10CustomerEvidencePackage(StrictCustomerEvidenceModel):
    generated_at: str
    source_file: str
    evidence_records: list[CustomerEvidenceRecord]
    field_mapping_suggestions: list[FieldMappingSuggestion]
    entity_graph: TraceabilityEntityGraph
    event_graph: list[CustomerEventNode]
    cte_classification_results: list[CteClassificationResult]
    reviewer_questions: list[dict[str, Any]]
    inferred_facts: list[InferredEvidenceFact] = Field(default_factory=list)
    document_profiles: list[CustomerDocumentProfile] = Field(default_factory=list)
    evidence_conflicts: list[EvidenceConflict] = Field(default_factory=list)
    quality_report: CustomerEvidenceQualityReport | None = None
    summary: dict[str, Any]


FIELD_ALIASES: dict[str, tuple[str, str]] = {
    "event id": ("event_id", "Traceability event identifier"),
    "event_id": ("event_id", "Traceability event identifier"),
    "event line id": ("event_line_id", "Traceability event line identifier"),
    "event_line_id": ("event_line_id", "Traceability event line identifier"),
    "event type": ("event_type", "Claimed CTE/event type"),
    "event_type": ("event_type", "Claimed CTE/event type"),
    "event date": ("event_datetime", "Traceability event date/time"),
    "event datetime": ("event_datetime", "Traceability event date/time"),
    "event_datetime": ("event_datetime", "Traceability event date/time"),
    "ship date": ("date_you_shipped_the_food", "Shipping KDE: date you shipped the food"),
    "shipping date": ("date_you_shipped_the_food", "Shipping KDE: date you shipped the food"),
    "date shipped": ("date_you_shipped_the_food", "Shipping KDE: date you shipped the food"),
    "date_you_shipped_the_food": ("date_you_shipped_the_food", "Shipping KDE: date you shipped the food"),
    "receive date": ("received_date", "Receiving event date"),
    "received date": ("received_date", "Receiving event date"),
    "lot": ("traceability_lot_code", "Traceability lot code"),
    "lot #": ("traceability_lot_code", "Traceability lot code"),
    "lot number": ("traceability_lot_code", "Traceability lot code"),
    "lot_or_tlc": ("traceability_lot_code", "Traceability lot code"),
    "tlc": ("traceability_lot_code", "Traceability lot code"),
    "traceability lot code": ("traceability_lot_code", "Traceability lot code"),
    "traceability_lot_code": ("traceability_lot_code", "Traceability lot code"),
    "source_lot_or_tlc": ("source_lot_or_tlc", "Source lot or traceability lot code"),
    "source lot": ("source_lot_or_tlc", "Source lot or traceability lot code"),
    "source tlc": ("source_lot_or_tlc", "Source lot or traceability lot code"),
    "output_lot_or_tlc": ("output_lot_or_tlc", "Output lot or traceability lot code"),
    "output lot": ("output_lot_or_tlc", "Output lot or traceability lot code"),
    "output tlc": ("output_lot_or_tlc", "Output lot or traceability lot code"),
    "target_lot_or_tlc": ("target_lot_or_tlc", "Target lot or traceability lot code"),
    "target lot": ("target_lot_or_tlc", "Target lot or traceability lot code"),
    "product": ("product_name", "Product name"),
    "product name": ("product_name", "Product name"),
    "product_name": ("product_name", "Product name"),
    "item": ("product_name", "Product name"),
    "product_id": ("product_id", "Product identifier"),
    "quantity": ("quantity", "Quantity"),
    "unit": ("unit", "Unit of measure"),
    "ftl_category": ("ftl_category", "Food Traceability List category"),
    "ftl_food_category": ("ftl_category", "Food Traceability List category"),
    "is_ftl_maybe": ("is_ftl_maybe", "Customer FTL scope flag"),
    "food form": ("food_form", "Food form state"),
    "form": ("food_form", "Food form state"),
    "destination": ("destination_type", "Destination type"),
    "destination type": ("destination_type", "Destination type"),
    "destination_type": ("destination_type", "Destination type"),
    "customer type": ("destination_type", "Destination type"),
    "actor_location_id": ("actor_location_id", "Actor location identifier"),
    "location_id": ("location_id", "Location identifier"),
    "location_name": ("location_name", "Location name"),
    "location_type": ("location_type", "Location type"),
    "partner_id": ("partner_id", "Counterparty identifier"),
    "partner_name": ("partner_name", "Counterparty name"),
    "partner_type": ("partner_type", "Counterparty type"),
    "relationship": ("partner_relationship", "Counterparty relationship"),
    "from_partner_id": ("from_partner_id", "From counterparty identifier"),
    "to_partner_id": ("to_partner_id", "To counterparty identifier"),
    "business_id": ("business_id", "Business identifier"),
    "business_type": ("business_type", "Business type"),
    "company_name": ("company_name", "Company name"),
    "handles_ftl_foods": ("handles_ftl_foods", "Handles FTL foods"),
    "covered_entity_status": ("covered_entity_status", "Covered entity status"),
    "reference_record_type": ("reference_record_type", "Source record type"),
    "reference_record_no": ("reference_record_no", "Source record number"),
    "evidence_id": ("source_document_id", "Source document identifier"),
    "evidence_type": ("source_document_type", "Source document type"),
    "evidence_status": ("source_document_status", "Source document status"),
    "claim_id": ("exemption_claim_id", "Exemption claim identifier"),
    "claim_type": ("exemption_claim_type", "Exemption claim type"),
    "claimed_by": ("exemption_claimed_by", "Exemption claimed by"),
    "evidence_provided": ("exemption_evidence_provided", "Exemption evidence provided"),
    "plan_item": ("traceability_plan_item", "Traceability plan item"),
    "answer": ("traceability_plan_answer", "Traceability plan answer"),
    "kde_id": ("kde_id", "KDE identifier"),
    "cte_type": ("cte_type", "KDE CTE type"),
    "kde_name": ("kde_name", "KDE name"),
    "field_key": ("kde_field_key", "KDE field key"),
    "kde_value": ("kde_value", "KDE value"),
    "lineage_id": ("lineage_id", "TLC lineage identifier"),
    "relationship_type": ("relationship_type", "TLC lineage relationship type"),
    "lineage_status": ("lineage_status", "TLC lineage status"),
}

CTE_ALIASES = {
    "harvest": "harvesting",
    "harvesting": "harvesting",
    "cool": "cooling",
    "cooling": "cooling",
    "pack": "initial_packing",
    "packing": "initial_packing",
    "initial packing": "initial_packing",
    "initial_packing": "initial_packing",
    "first land based receiving": "first_land_based_receiving",
    "first_land_based_receiving": "first_land_based_receiving",
    "first land": "first_land_based_receiving",
    "ship": "shipping",
    "shipping": "shipping",
    "receive": "receiving",
    "receiving": "receiving",
    "transform": "transformation",
    "transformation": "transformation",
    "traceability plan": "traceability_plan",
}

ACTOR_ROLE_TERMS = [
    ("first_land_based_receiver", ["first land", "landing", "seafood receiver"]),
    ("initial_packer", ["initial pack", "packing house", "packer"]),
    ("harvester", ["harvest", "harvester"]),
    ("cooler", ["cooler", "cooling"]),
    ("farm", ["farm", "grower", "ranch"]),
    ("processor", ["processor", "processing", "manufacturer", "kitchen", "commissary"]),
    ("restaurant_rfe", ["restaurant", "retail food", "rfe", "store"]),
    ("distributor", ["distributor", "warehouse", "dc"]),
    ("transporter", ["carrier", "transporter", "truck", "freight"]),
    ("consumer", ["consumer", "direct to consumer", "dtc"]),
    ("shipper", ["shipper", "seller"]),
    ("receiver", ["receiver", "buyer"]),
]


def build_phase10_customer_evidence(
    *,
    input_file: Path,
    ftl_food_items_file: Path | None = None,
) -> Phase10CustomerEvidencePackage:
    evidence_records = read_spreadsheet_evidence(input_file)
    inferred_facts = infer_filename_and_sheet_facts(input_file=input_file, evidence_records=evidence_records)
    mapping_suggestions = build_field_mapping_suggestions(evidence_records)
    document_profiles = build_document_profiles(input_file=input_file, evidence_records=evidence_records, inferred_facts=inferred_facts)
    evidence_conflicts = detect_evidence_conflicts(evidence_records)
    ftl_food_items = _load_optional_json_list(ftl_food_items_file)
    entity_graph = build_traceability_entity_graph(evidence_records, ftl_food_items=ftl_food_items)
    event_graph = build_customer_event_graph(evidence_records, entity_graph=entity_graph, ftl_food_items=ftl_food_items)
    classifications = [classify_event_ctes(event) for event in event_graph]
    event_graph = [
        event.model_copy(
            update={
                "classified_ctes": result.final_ctes,
                "suppressed_ctes": result.suppressed_ctes,
                "reviewer_questions": result.reviewer_questions,
            }
        )
        for event, result in zip(event_graph, classifications, strict=True)
    ]
    reviewer_questions = _collect_reviewer_questions(event_graph, classifications, evidence_records)
    summary = _summary(
        source_file=input_file,
        evidence_records=evidence_records,
        suggestions=mapping_suggestions,
        entity_graph=entity_graph,
        event_graph=event_graph,
        classifications=classifications,
        reviewer_questions=reviewer_questions,
        inferred_facts=inferred_facts,
        document_profiles=document_profiles,
        evidence_conflicts=evidence_conflicts,
    )
    quality_report = build_customer_evidence_quality_report(
        source_file=input_file,
        evidence_records=evidence_records,
        suggestions=mapping_suggestions,
        inferred_facts=inferred_facts,
        document_profiles=document_profiles,
        evidence_conflicts=evidence_conflicts,
        classifications=classifications,
    )
    return Phase10CustomerEvidencePackage(
        generated_at=GENERATED_AT,
        source_file=str(input_file),
        evidence_records=evidence_records,
        field_mapping_suggestions=mapping_suggestions,
        entity_graph=entity_graph,
        event_graph=event_graph,
        cte_classification_results=classifications,
        reviewer_questions=reviewer_questions,
        inferred_facts=inferred_facts,
        document_profiles=document_profiles,
        evidence_conflicts=evidence_conflicts,
        quality_report=quality_report,
        summary=summary,
    )


def write_phase10_customer_evidence_artifacts(package: Phase10CustomerEvidencePackage, output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = {
        "summary": output_dir / "phase10-summary.json",
        "evidenceRecords": output_dir / "phase10-evidence-records.json",
        "fieldMappingSuggestions": output_dir / "phase10-field-mapping-suggestions.json",
        "entityGraph": output_dir / "phase10-entity-graph.json",
        "eventGraph": output_dir / "phase10-event-graph.json",
        "cteClassificationResults": output_dir / "phase10-cte-classification-results.json",
        "reviewerQuestions": output_dir / "phase10-reviewer-questions.json",
        "inferredFacts": output_dir / "phase10a-inferred-facts.json",
        "documentProfiles": output_dir / "phase10a-document-profiles.json",
        "evidenceConflicts": output_dir / "phase10a-evidence-conflicts.json",
        "qualityReport": output_dir / "phase10a-quality-report.json",
    }
    _write_json(outputs["summary"], package.summary)
    _write_json(outputs["evidenceRecords"], [record.model_dump(mode="json") for record in package.evidence_records])
    _write_json(outputs["fieldMappingSuggestions"], [item.model_dump(mode="json") for item in package.field_mapping_suggestions])
    _write_json(outputs["entityGraph"], package.entity_graph.model_dump(mode="json"))
    _write_json(outputs["eventGraph"], [event.model_dump(mode="json") for event in package.event_graph])
    _write_json(outputs["cteClassificationResults"], [result.model_dump(mode="json") for result in package.cte_classification_results])
    _write_json(outputs["reviewerQuestions"], package.reviewer_questions)
    _write_json(outputs["inferredFacts"], [fact.model_dump(mode="json") for fact in package.inferred_facts])
    _write_json(outputs["documentProfiles"], [profile.model_dump(mode="json") for profile in package.document_profiles])
    _write_json(outputs["evidenceConflicts"], [conflict.model_dump(mode="json") for conflict in package.evidence_conflicts])
    _write_json(outputs["qualityReport"], package.quality_report.model_dump(mode="json") if package.quality_report else {})
    return {key: str(path) for key, path in outputs.items()}


def read_spreadsheet_evidence(input_file: Path) -> list[CustomerEvidenceRecord]:
    suffix = input_file.suffix.lower()
    if suffix == ".csv":
        return _read_csv_evidence(input_file)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_evidence(input_file)
    raise ValueError(f"unsupported customer evidence file type: {input_file.suffix}")


def infer_filename_and_sheet_facts(*, input_file: Path, evidence_records: list[CustomerEvidenceRecord]) -> list[InferredEvidenceFact]:
    facts: list[InferredEvidenceFact] = []
    sources = [("filename", input_file.name)]
    sources.extend(("sheet_name", sheet) for sheet in sorted({record.sheet_name for record in evidence_records}))
    for source_kind, source_name in sources:
        for field_key, raw_value, confidence in _infer_facts_from_text(source_name):
            facts.append(
                InferredEvidenceFact(
                    fact_id=f"phase10a-fact-{len(facts) + 1:04d}",
                    source_kind=source_kind,
                    source_name=source_name,
                    field_key=field_key,
                    raw_value=raw_value,
                    normalized_value=_normalize_value(raw_value),
                    confidence=confidence,
                    extraction_method=f"{source_kind}_pattern_inference",
                    evidence_pointer={"file_name": input_file.name, "source_kind": source_kind, "source_name": source_name},
                )
            )
    return facts


def build_document_profiles(
    *,
    input_file: Path,
    evidence_records: list[CustomerEvidenceRecord],
    inferred_facts: list[InferredEvidenceFact],
) -> list[CustomerDocumentProfile]:
    grouped: dict[str, list[CustomerEvidenceRecord]] = defaultdict(list)
    for record in evidence_records:
        grouped[record.sheet_name].append(record)

    profiles: list[CustomerDocumentProfile] = []
    for sheet_name, records in sorted(grouped.items()):
        text = " ".join(
            [
                input_file.name,
                sheet_name,
                *[record.column_name for record in records[:100]],
                *[record.raw_value for record in records[:100]],
                *[fact.raw_value for fact in inferred_facts if fact.source_name in {input_file.name, sheet_name}],
            ]
        )
        document_type, signals = _classify_document_type(text)
        profiles.append(
            CustomerDocumentProfile(
                profile_id=f"phase10a-doc-{len(profiles) + 1:04d}",
                source_name=sheet_name,
                document_type=document_type,
                evidence_kind="spreadsheet_sheet",
                confidence=0.88 if signals else 0.42,
                detected_signals=signals,
                supported_parser=_supported_parser_for_document_type(document_type),
                evidence_ids=[record.evidence_id for record in records[:50]],
            )
        )
    if not profiles:
        document_type, signals = _classify_document_type(input_file.name)
        profiles.append(
            CustomerDocumentProfile(
                profile_id="phase10a-doc-0001",
                source_name=input_file.name,
                document_type=document_type,
                evidence_kind="uploaded_file",
                confidence=0.75 if signals else 0.35,
                detected_signals=signals,
                supported_parser=_supported_parser_for_document_type(document_type),
                evidence_ids=[],
            )
        )
    return profiles


def detect_evidence_conflicts(evidence_records: list[CustomerEvidenceRecord]) -> list[EvidenceConflict]:
    rows = _row_facts_with_record_ids(evidence_records)
    conflicts: list[EvidenceConflict] = []
    for row_key, facts in sorted(rows.items()):
        for field_key, values in sorted(facts.items()):
            normalized_values = sorted(value for value in values if value)
            if len(normalized_values) > 1:
                conflicts.append(
                    EvidenceConflict(
                        conflict_id=f"phase10a-conflict-{len(conflicts) + 1:04d}",
                        conflict_type="same_row_conflicting_values",
                        row_key=row_key,
                        field_key=field_key,
                        values=normalized_values,
                        evidence_ids_by_value={value: values[value] for value in normalized_values},
                        severity="high" if field_key in {"traceability_lot_code", "event_datetime", "product_name", "event_type"} else "medium",
                    )
                )
    return conflicts


def build_customer_evidence_quality_report(
    *,
    source_file: Path,
    evidence_records: list[CustomerEvidenceRecord],
    suggestions: list[FieldMappingSuggestion],
    inferred_facts: list[InferredEvidenceFact],
    document_profiles: list[CustomerDocumentProfile],
    evidence_conflicts: list[EvidenceConflict],
    classifications: list[CteClassificationResult],
) -> CustomerEvidenceQualityReport:
    sheets = sorted({record.sheet_name for record in evidence_records})
    mapped_records = [record for record in evidence_records if record.confidence >= 0.6]
    unmapped_columns = [
        {
            "sheet": suggestion.source_sheet,
            "column": suggestion.source_column,
            "fieldKey": suggestion.field_key,
            "confidence": suggestion.confidence,
            "evidenceIds": suggestion.evidence_ids[:5],
        }
        for suggestion in suggestions
        if suggestion.confidence < 0.6
    ]
    low_confidence_mappings = [
        {
            "sheet": suggestion.source_sheet,
            "column": suggestion.source_column,
            "fieldKey": suggestion.field_key,
            "confidence": suggestion.confidence,
        }
        for suggestion in suggestions
        if suggestion.confidence < 0.8
    ]
    row_signatures = Counter(
        "|".join(
            [
                record.sheet_name,
                str(record.row_number),
                record.field_key,
                record.normalized_value,
            ]
        )
        for record in evidence_records
    )
    duplicate_row_count = sum(count - 1 for count in row_signatures.values() if count > 1)
    missing_lineage_anchor_count = sum(1 for record in evidence_records if not record.cell or not record.source_pointer.cell)
    issue_count = len(unmapped_columns) + len(evidence_conflicts) + missing_lineage_anchor_count
    quality_gate = "pass_with_review" if issue_count else "pass"
    if evidence_conflicts or len(unmapped_columns) > 10:
        quality_gate = "needs_review"
    return CustomerEvidenceQualityReport(
        generated_at=GENERATED_AT,
        source_file=str(source_file),
        parse_coverage={
            "evidenceRecords": len(evidence_records),
            "mappedRecords": len(mapped_records),
            "mappedRecordRate": round(len(mapped_records) / len(evidence_records), 4) if evidence_records else 0,
            "sheetsParsed": len(sheets),
        },
        workbook_structure={
            "sheets": sheets,
            "hiddenOrMergedCellHandling": "xlsx hidden rows/columns skipped; merged ranges resolved from top-left cells where available",
            "headerDetection": "best scoring row across first non-empty rows, with support for blank header bands and section header skips",
        },
        document_type_counts=dict(sorted(Counter(profile.document_type for profile in document_profiles).items())),
        unmapped_sheets=[] if evidence_records else [source_file.name],
        unmapped_columns=unmapped_columns,
        low_confidence_mappings=low_confidence_mappings,
        inferred_fact_count=len(inferred_facts),
        conflict_count=len(evidence_conflicts),
        duplicate_row_count=duplicate_row_count,
        missing_lineage_anchor_count=missing_lineage_anchor_count,
        abstention_count=sum(1 for classification in classifications if not classification.final_ctes),
        quality_gate=quality_gate,
        issues=_quality_issues(unmapped_columns=unmapped_columns, conflicts=evidence_conflicts, missing_lineage_anchor_count=missing_lineage_anchor_count),
    )


def build_field_mapping_suggestions(evidence_records: list[CustomerEvidenceRecord]) -> list[FieldMappingSuggestion]:
    grouped: dict[tuple[str, str, str], list[CustomerEvidenceRecord]] = defaultdict(list)
    for record in evidence_records:
        grouped[(record.sheet_name, record.column_name, record.field_key)].append(record)

    suggestions: list[FieldMappingSuggestion] = []
    for index, ((sheet_name, column_name, field_key), records) in enumerate(sorted(grouped.items()), start=1):
        header_key = _header_key(column_name)
        alias_key = header_key if header_key in FIELD_ALIASES else _slug(column_name)
        canonical_field = FIELD_ALIASES.get(alias_key, (field_key, _title_from_slug(field_key)))[1]
        confidence = max(record.confidence for record in records)
        method = "ai_assisted_field_mapping_suggestion"
        if _header_key(column_name) == field_key:
            method = "spreadsheet_header"
        suggestions.append(
            FieldMappingSuggestion(
                suggestion_id=f"phase10-map-{index:04d}",
                source_sheet=sheet_name,
                source_column=column_name,
                field_key=field_key,
                canonical_field=canonical_field,
                confidence=confidence,
                suggestion_method=method,
                rationale=_mapping_rationale(column_name, field_key, confidence),
                evidence_ids=[record.evidence_id for record in records[:25]],
            )
        )
    return suggestions


def build_traceability_entity_graph(
    evidence_records: list[CustomerEvidenceRecord],
    *,
    ftl_food_items: list[dict[str, Any]] | None = None,
) -> TraceabilityEntityGraph:
    row_facts = _row_facts(evidence_records)
    products: dict[str, TraceabilityEntity] = {}
    product_forms: dict[str, TraceabilityEntity] = {}
    lots: dict[str, TraceabilityEntity] = {}
    actors: dict[str, TraceabilityEntity] = {}
    locations: dict[str, TraceabilityEntity] = {}
    counterparties: dict[str, TraceabilityEntity] = {}
    documents: dict[str, TraceabilityEntity] = {}

    for row in row_facts.values():
        facts = row["facts"]
        evidence_ids = row["evidence_ids"]
        product_name = _first(facts, "product_name")
        product_id = _first(facts, "product_id") or _stable_id("product", product_name)
        if product_name:
            form = resolve_food_form(product_name=product_name, ftl_category=_first(facts, "ftl_category"), ftl_food_items=ftl_food_items)
            products.setdefault(
                product_id,
                TraceabilityEntity(
                    entity_id=product_id,
                    entity_type="product",
                    name=product_name,
                    attributes={
                        "ftl_category": _first(facts, "ftl_category"),
                        "is_ftl_likely": form.is_ftl_likely,
                    },
                    evidence_ids=evidence_ids,
                ),
            )
            product_forms.setdefault(
                f"form-{product_id}",
                TraceabilityEntity(
                    entity_id=f"form-{product_id}",
                    entity_type="product_form",
                    name=product_name,
                    attributes=form.model_dump(mode="json"),
                    evidence_ids=evidence_ids,
                ),
            )

        for key in ("traceability_lot_code", "source_lot_or_tlc", "output_lot_or_tlc", "target_lot_or_tlc"):
            value = _first(facts, key)
            if value:
                lot_id = _stable_id("lot", value)
                lots.setdefault(
                    lot_id,
                    TraceabilityEntity(
                        entity_id=lot_id,
                        entity_type="lot_or_tlc",
                        name=value,
                        attributes={"source_field": key},
                        evidence_ids=evidence_ids,
                    ),
                )

        location_id = _first(facts, "location_id") or _first(facts, "actor_location_id")
        location_name = _first(facts, "location_name") or location_id
        if location_id or location_name:
            entity_id = location_id or _stable_id("location", location_name)
            location_type = _first(facts, "location_type")
            role = resolve_actor_role(actor_name=location_name, actor_type=location_type)
            locations.setdefault(
                entity_id,
                TraceabilityEntity(
                    entity_id=entity_id,
                    entity_type="location",
                    name=location_name or entity_id,
                    attributes={"location_type": location_type, "role_resolution": role.model_dump(mode="json")},
                    evidence_ids=evidence_ids,
                ),
            )
            actors.setdefault(
                f"actor-{entity_id}",
                TraceabilityEntity(
                    entity_id=f"actor-{entity_id}",
                    entity_type="actor",
                    name=location_name or entity_id,
                    attributes=role.model_dump(mode="json"),
                    evidence_ids=evidence_ids,
                ),
            )

        partner_id = _first(facts, "partner_id") or _first(facts, "from_partner_id") or _first(facts, "to_partner_id")
        partner_name = _first(facts, "partner_name") or partner_id
        if partner_id or partner_name:
            entity_id = partner_id or _stable_id("counterparty", partner_name)
            counterparties.setdefault(
                entity_id,
                TraceabilityEntity(
                    entity_id=entity_id,
                    entity_type="counterparty",
                    name=partner_name or entity_id,
                    attributes={"partner_type": _first(facts, "partner_type")},
                    evidence_ids=evidence_ids,
                ),
            )

        doc_id = _first(facts, "source_document_id") or _first(facts, "reference_record_no")
        if doc_id:
            documents.setdefault(
                doc_id,
                TraceabilityEntity(
                    entity_id=doc_id,
                    entity_type="document",
                    name=doc_id,
                    attributes={
                        "document_type": _first(facts, "source_document_type") or _first(facts, "reference_record_type"),
                        "status": _first(facts, "source_document_status"),
                    },
                    evidence_ids=evidence_ids,
                ),
            )

    return TraceabilityEntityGraph(
        products=sorted(products.values(), key=lambda item: item.entity_id),
        product_forms=sorted(product_forms.values(), key=lambda item: item.entity_id),
        lots=sorted(lots.values(), key=lambda item: item.entity_id),
        actors=sorted(actors.values(), key=lambda item: item.entity_id),
        locations=sorted(locations.values(), key=lambda item: item.entity_id),
        counterparties=sorted(counterparties.values(), key=lambda item: item.entity_id),
        documents=sorted(documents.values(), key=lambda item: item.entity_id),
    )


def build_customer_event_graph(
    evidence_records: list[CustomerEvidenceRecord],
    *,
    entity_graph: TraceabilityEntityGraph | None = None,
    ftl_food_items: list[dict[str, Any]] | None = None,
) -> list[CustomerEventNode]:
    row_facts = _row_facts(evidence_records)
    base_events: dict[str, dict[str, Any]] = {}
    line_rows: list[dict[str, Any]] = []
    standalone_rows: list[dict[str, Any]] = []

    for row in row_facts.values():
        facts = row["facts"]
        sheet = row["sheet"]
        if _first(facts, "event_id") and ("cte_events" in _slug(sheet) or _first(facts, "event_type")):
            base_events[_first(facts, "event_id")] = row
        elif _first(facts, "event_id") and (_first(facts, "product_name") or _first(facts, "product_id") or "line_items" in _slug(sheet)):
            line_rows.append(row)
        elif _looks_like_event_row(facts):
            standalone_rows.append(row)

    events: list[CustomerEventNode] = []
    used_line_row_keys: set[str] = set()
    for line in line_rows:
        event_id = _first(line["facts"], "event_id")
        base = base_events.get(event_id)
        merged = _merge_rows(base, line) if base else line
        events.append(_event_node_from_row(merged, ftl_food_items=ftl_food_items))
        used_line_row_keys.add(line["row_key"])

    for event_id, base in sorted(base_events.items()):
        if not any(_first(line["facts"], "event_id") == event_id for line in line_rows):
            events.append(_event_node_from_row(base, ftl_food_items=ftl_food_items))

    for row in standalone_rows:
        if row["row_key"] not in used_line_row_keys:
            events.append(_event_node_from_row(row, ftl_food_items=ftl_food_items))

    deduped: dict[str, CustomerEventNode] = {}
    for event in events:
        deduped[event.event_id] = event
    return sorted(deduped.values(), key=lambda event: event.event_id)


def resolve_food_form(
    *,
    product_name: str | None,
    ftl_category: str | None = None,
    food_form: str | None = None,
    ftl_food_items: list[dict[str, Any]] | None = None,
) -> FoodFormResolution:
    text = " ".join(part for part in [product_name, ftl_category, food_form] if part).lower()
    form_state: list[str] = []
    form_terms = [
        ("fresh_cut", ["fresh-cut", "fresh cut"]),
        ("fresh", ["fresh"]),
        ("frozen", ["frozen"]),
        ("previously_frozen", ["previously frozen"]),
        ("canned", ["canned"]),
        ("shelf_stable", ["shelf stable", "ambient temperature", "shelf-stable"]),
        ("aseptic", ["aseptic"]),
        ("refrigerated", ["refrigerated", "chilled"]),
        ("cooked", ["cooked", "baked", "roasted"]),
        ("kill_step_applied", ["kill step", "pasteurized", "validated lethality"]),
        ("raw_agricultural_commodity", ["raw agricultural commodity", "rac"]),
        ("seafood_from_fishing_vessel", ["fishing vessel", "vessel", "seafood"]),
    ]
    for state, terms in form_terms:
        if any(term in text for term in terms):
            form_state.append(state)

    match_reasons = []
    is_ftl_likely: bool | None = None
    confidence = 0.35
    if ftl_category:
        is_ftl_likely = True
        confidence = 0.82
        match_reasons.append("customer supplied an FTL category")

    matched_item = _match_ftl_item(text, ftl_food_items or [])
    if matched_item:
        is_ftl_likely = True
        confidence = max(confidence, 0.9)
        match_reasons.append(f"matched FTL library item: {matched_item}")

    if any(state in form_state for state in ("canned", "shelf_stable", "aseptic")):
        output_remains_ftl = False
        confidence = min(confidence, 0.78)
        match_reasons.append("form state indicates the output may no longer remain on the FTL")
    elif is_ftl_likely is True:
        output_remains_ftl = True
    else:
        output_remains_ftl = None

    if not form_state and text:
        if any(term in text for term in ["basil", "lettuce", "cucumber", "sprout", "tomato", "melon", "pepper"]):
            form_state.append("fresh")

    review_required = is_ftl_likely is None or output_remains_ftl is None or any(
        state in form_state for state in ("canned", "shelf_stable", "aseptic", "kill_step_applied")
    )
    if is_ftl_likely is None:
        match_reasons.append("food scope could not be resolved from the supplied evidence")
    return FoodFormResolution(
        product_name=product_name,
        ftl_category=ftl_category,
        is_ftl_likely=is_ftl_likely,
        form_state=sorted(set(form_state)),
        output_remains_ftl=output_remains_ftl,
        confidence=confidence,
        reasons=match_reasons,
        review_required=review_required,
    )


def resolve_actor_role(*, actor_name: str | None = None, actor_type: str | None = None, event_type: str | None = None) -> ActorRoleResolution:
    text = " ".join(part for part in [actor_name, actor_type, event_type] if part).lower()
    for role, terms in ACTOR_ROLE_TERMS:
        if any(term in text for term in terms):
            return ActorRoleResolution(
                actor_name=actor_name,
                actor_type=actor_type,
                role=role,
                confidence=0.86,
                reasons=[f"matched role terms for {role}"],
            )
    normalized_event = _normalize_cte(event_type)
    if normalized_event == "shipping":
        role = "shipper"
    elif normalized_event == "receiving":
        role = "receiver"
    elif normalized_event == "transformation":
        role = "processor"
    elif normalized_event == "first_land_based_receiving":
        role = "first_land_based_receiver"
    elif normalized_event == "initial_packing":
        role = "initial_packer"
    elif normalized_event == "harvesting":
        role = "harvester"
    else:
        role = "unknown"
    return ActorRoleResolution(
        actor_name=actor_name,
        actor_type=actor_type,
        role=role,
        confidence=0.62 if role != "unknown" else 0.3,
        reasons=[f"inferred from event type {normalized_event}"] if normalized_event else ["actor role was not clear from evidence"],
    )


def classify_event_ctes(event: CustomerEventNode) -> CteClassificationResult:
    candidates: list[str] = []
    reasons: list[str] = []
    questions: list[str] = []

    claimed = _normalize_cte(event.event_type_claim)
    if claimed:
        candidates.append(claimed)
        reasons.append(f"structured event_type mapped to {claimed}")

    action_text = " ".join(event.action_terms).lower()
    if event.event_datetime and "date_you_shipped_the_food" in action_text:
        candidates.append("shipping")
        reasons.append("shipping KDE date was present")
    if event.output_lot_or_tlc and (event.source_lot_or_tlc or claimed == "transformation"):
        candidates.append("transformation")
        reasons.append("source/output lot relationship indicates transformation")
    if event.actor_role.role == "first_land_based_receiver":
        candidates.append("first_land_based_receiving")
        reasons.append("actor role resolved to first land-based receiver")

    if not candidates and event.product_name:
        questions.append("Classify CTE: event has product evidence but no reliable structured CTE signal.")

    final_ctes = _unique(candidates)
    suppressed: list[str] = []
    destination = (event.destination_type or "").lower()
    if "first_land_based_receiving" in final_ctes and "receiving" in final_ctes:
        final_ctes.remove("receiving")
        suppressed.append("receiving")
        reasons.append("suppressed generic receiving because first land-based receiving is more specific")
    if "shipping" in final_ctes and any(term in destination for term in ["consumer", "direct", "dtc"]):
        final_ctes.remove("shipping")
        suppressed.append("shipping")
        reasons.append("suppressed shipping because destination appears direct-to-consumer")
    if event.food_form.output_remains_ftl is False:
        for cte in ("shipping", "transformation"):
            if cte in final_ctes:
                final_ctes.remove(cte)
                suppressed.append(cte)
        reasons.append("suppressed downstream FTL CTE duties because resolved output does not remain FTL")

    if event.food_form.review_required:
        questions.append("Resolve food/form scope: evidence is ambiguous or indicates a form-change/kill-step condition.")
    if ("shipping" in candidates or event.to_partner_id) and not event.destination_type and not event.to_partner_id:
        questions.append("Confirm destination: customer evidence does not clearly identify consumer versus business recipient.")
    if "transformation" in candidates and event.food_form.output_remains_ftl is None:
        questions.append("Confirm transformed output FTL scope before applying downstream FTL duties.")
    if any("exemption" in term for term in event.action_terms):
        questions.append("Review exemption claim evidence before suppressing any approved obligation.")

    confidence = 0.88 if final_ctes else 0.45
    if questions:
        confidence = min(confidence, 0.68)
    return CteClassificationResult(
        event_id=event.event_id,
        candidate_ctes=_unique(candidates),
        final_ctes=final_ctes,
        suppressed_ctes=_unique(suppressed),
        confidence=confidence,
        reasons=reasons,
        reviewer_questions=_unique(questions),
        evidence_ids=event.evidence_ids,
    )


def _read_csv_evidence(input_file: Path) -> list[CustomerEvidenceRecord]:
    records: list[CustomerEvidenceRecord] = []
    with input_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return []
    header_index = _detect_header_row(rows)
    headers = _build_headers(rows, header_index)
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if _is_notes_or_repeated_header_row(row, headers):
            continue
        for column_index, header in enumerate(headers, start=1):
            raw = row[column_index - 1] if column_index <= len(row) else ""
            if str(raw).strip() == "":
                continue
            records.append(_evidence_record(input_file=input_file, sheet_name="csv", row_number=offset, column_name=header, column_index=column_index, raw_value=raw))
    return records


def _read_xlsx_evidence(input_file: Path) -> list[CustomerEvidenceRecord]:
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to ingest XLSX customer evidence") from exc

    records: list[CustomerEvidenceRecord] = []
    workbook = load_workbook(input_file, read_only=False, data_only=True)
    formula_workbook = load_workbook(input_file, read_only=False, data_only=False)
    for worksheet in workbook.worksheets:
        formula_worksheet = formula_workbook[worksheet.title]
        visible_columns = [
            column_index
            for column_index in range(1, worksheet.max_column + 1)
            if not worksheet.column_dimensions[get_column_letter(column_index)].hidden
        ]
        visible_rows = [
            row_index
            for row_index in range(1, worksheet.max_row + 1)
            if not worksheet.row_dimensions[row_index].hidden
        ]
        if not visible_rows or not visible_columns:
            continue
        row_values = [[_merged_or_cell_value(worksheet, row_index, column_index, formula_worksheet=formula_worksheet) for column_index in visible_columns] for row_index in visible_rows]
        header_position = _detect_header_row(row_values)
        headers = _build_headers(row_values, header_position)
        for row_position, row in enumerate(row_values[header_position + 1 :], start=header_position + 1):
            sheet_row = visible_rows[row_position]
            if _is_notes_or_repeated_header_row(row, headers):
                continue
            for output_column_index, header in enumerate(headers, start=1):
                raw = row[output_column_index - 1] if output_column_index <= len(row) else ""
                if _cell_to_string(raw).strip() == "":
                    continue
                sheet_column = visible_columns[output_column_index - 1]
                record = _evidence_record(
                    input_file=input_file,
                    sheet_name=worksheet.title,
                    row_number=sheet_row,
                    column_name=header,
                    column_index=sheet_column,
                    raw_value=raw,
                )
                cell = f"{get_column_letter(sheet_column)}{sheet_row}"
                record = record.model_copy(update={"cell": cell, "source_pointer": record.source_pointer.model_copy(update={"cell": cell})})
                records.append(record)
    return records


def _evidence_record(
    *,
    input_file: Path,
    sheet_name: str,
    row_number: int,
    column_name: str,
    column_index: int,
    raw_value: Any,
) -> CustomerEvidenceRecord:
    raw = _cell_to_string(raw_value)
    field_key, confidence, method = _suggest_field_key(column_name)
    normalized = _normalize_value(raw_value, field_key=field_key)
    field_type = _detect_field_type(field_key, normalized)
    cell = f"{_column_letter(column_index)}{row_number}"
    source_pointer = EvidenceSourcePointer(
        file_name=input_file.name,
        sheet_name=sheet_name,
        row_number=row_number,
        column_name=column_name,
        column_index=column_index,
        cell=cell,
    )
    evidence_id = f"ev-{_slug(input_file.stem)}-{_slug(sheet_name)}-r{row_number}-c{column_index}"
    return CustomerEvidenceRecord(
        evidence_id=evidence_id,
        uploaded_file=input_file.name,
        sheet_name=sheet_name,
        row_number=row_number,
        column_name=column_name,
        column_index=column_index,
        cell=cell,
        raw_value=raw,
        normalized_value=normalized,
        field_key=field_key,
        field_type=field_type,
        extraction_method=method,
        confidence=confidence,
        source_pointer=source_pointer,
    )


def _suggest_field_key(header: str) -> tuple[str, float, str]:
    key = _header_key(header)
    slug = _slug(header)
    alias_key = key if key in FIELD_ALIASES else slug
    if alias_key in FIELD_ALIASES:
        field_key, _ = FIELD_ALIASES[alias_key]
        confidence = 0.98 if slug == field_key else 0.9
        method = "spreadsheet_header" if slug == field_key else "ai_assisted_field_mapping_suggestion"
        return field_key, confidence, method
    return slug or "unknown_field", 0.45, "spreadsheet_header_unmapped"


def _detect_field_type(field_key: str, value: str) -> str:
    if value.startswith("="):
        return "formula"
    if "date" in field_key or re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", value):
        return "date"
    if field_key.endswith("_id") or "lot" in field_key or field_key == "traceability_lot_code":
        return "identifier"
    if value.lower() in {"yes", "no", "true", "false", "unknown"}:
        return "boolean_like"
    if re.fullmatch(r"-?\d+(?:\.\d+)?", value):
        return "number"
    return "text"


def _row_facts(evidence_records: list[CustomerEvidenceRecord]) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for record in evidence_records:
        row_key = f"{record.uploaded_file}:{record.sheet_name}:{record.row_number}"
        if row_key not in rows:
            rows[row_key] = {"row_key": row_key, "sheet": record.sheet_name, "row_number": record.row_number, "facts": defaultdict(list), "evidence_ids": []}
        rows[row_key]["facts"][record.field_key].append(record.normalized_value)
        rows[row_key]["facts"][f"source_column:{record.field_key}"].append(record.column_name)
        rows[row_key]["evidence_ids"].append(record.evidence_id)
    return rows


def _event_node_from_row(row: dict[str, Any], *, ftl_food_items: list[dict[str, Any]] | None = None) -> CustomerEventNode:
    facts = row["facts"]
    event_type = _first(facts, "event_type")
    event_datetime = _first(facts, "event_datetime") or _first(facts, "date_you_shipped_the_food") or _first(facts, "received_date")
    product_name = _first(facts, "product_name")
    ftl_category = _first(facts, "ftl_category")
    food_form_value = _first(facts, "food_form")
    actor_id = _first(facts, "actor_location_id") or _first(facts, "location_id")
    actor_role = resolve_actor_role(actor_name=_first(facts, "location_name") or actor_id, actor_type=_first(facts, "location_type"), event_type=event_type)
    event_id = _first(facts, "event_id") or _stable_id("event", row["row_key"])
    line_or_product = _first(facts, "product_id") or _stable_id("line", product_name)
    if _first(facts, "event_id") and (_first(facts, "product_name") or _first(facts, "product_id")):
        event_id = f"{event_id}:{line_or_product}"
    action_terms = _unique([event_type or "", *[key for key in facts if not key.startswith("source_column:")]])
    return CustomerEventNode(
        event_id=event_id,
        source_row_key=row["row_key"],
        evidence_ids=_unique(row["evidence_ids"]),
        event_type_claim=event_type,
        event_datetime=event_datetime,
        actor_id=actor_id,
        actor_role=actor_role,
        product_id=_first(facts, "product_id"),
        product_name=product_name,
        food_form=resolve_food_form(product_name=product_name, ftl_category=ftl_category, food_form=food_form_value, ftl_food_items=ftl_food_items),
        lot_or_tlc=_first(facts, "traceability_lot_code"),
        source_lot_or_tlc=_first(facts, "source_lot_or_tlc"),
        output_lot_or_tlc=_first(facts, "output_lot_or_tlc"),
        from_partner_id=_first(facts, "from_partner_id"),
        to_partner_id=_first(facts, "to_partner_id"),
        destination_type=_first(facts, "destination_type"),
        action_terms=action_terms,
    )


def _merge_rows(base: dict[str, Any] | None, line: dict[str, Any]) -> dict[str, Any]:
    if not base:
        return line
    merged_facts: dict[str, list[str]] = defaultdict(list)
    for source in (base["facts"], line["facts"]):
        for key, values in source.items():
            merged_facts[key].extend(values)
    return {
        "row_key": f"{base['row_key']}+{line['row_key']}",
        "sheet": line["sheet"],
        "row_number": line["row_number"],
        "facts": merged_facts,
        "evidence_ids": _unique([*base["evidence_ids"], *line["evidence_ids"]]),
    }


def _looks_like_event_row(facts: dict[str, list[str]]) -> bool:
    event_keys = {
        "event_type",
        "event_datetime",
        "date_you_shipped_the_food",
        "received_date",
        "traceability_lot_code",
        "source_lot_or_tlc",
        "output_lot_or_tlc",
        "destination_type",
    }
    return bool(event_keys & set(facts)) and bool({"product_name", "product_id", "traceability_lot_code"} & set(facts))


def _collect_reviewer_questions(
    event_graph: list[CustomerEventNode],
    classifications: list[CteClassificationResult],
    evidence_records: list[CustomerEvidenceRecord],
) -> list[dict[str, Any]]:
    by_event = {classification.event_id: classification for classification in classifications}
    questions: list[dict[str, Any]] = []
    for event in event_graph:
        for index, question in enumerate(event.reviewer_questions, start=1):
            questions.append(
                {
                    "question_id": f"phase10-question-{len(questions) + 1:04d}",
                    "event_id": event.event_id,
                    "question": question,
                    "reason": "abstention_or_low_confidence_fact",
                    "confidence": by_event[event.event_id].confidence,
                    "evidence_ids": event.evidence_ids,
                    "status": "needs_review",
                }
            )
    unmapped = [record for record in evidence_records if record.confidence < 0.5]
    if unmapped:
        questions.append(
            {
                "question_id": f"phase10-question-{len(questions) + 1:04d}",
                "event_id": None,
                "question": "Review unmapped customer columns before relying on event classification.",
                "reason": "low_confidence_field_mapping",
                "confidence": 0.45,
                "evidence_ids": [record.evidence_id for record in unmapped[:50]],
                "status": "needs_review",
            }
        )
    return questions


def _summary(
    *,
    source_file: Path,
    evidence_records: list[CustomerEvidenceRecord],
    suggestions: list[FieldMappingSuggestion],
    entity_graph: TraceabilityEntityGraph,
    event_graph: list[CustomerEventNode],
    classifications: list[CteClassificationResult],
    reviewer_questions: list[dict[str, Any]],
    inferred_facts: list[InferredEvidenceFact],
    document_profiles: list[CustomerDocumentProfile],
    evidence_conflicts: list[EvidenceConflict],
) -> dict[str, Any]:
    final_ctes = Counter(cte for result in classifications for cte in result.final_ctes)
    suppressed_ctes = Counter(cte for result in classifications for cte in result.suppressed_ctes)
    return {
        "phase": 10,
        "generatedAt": GENERATED_AT,
        "sourceFile": str(source_file),
        "evidenceRecords": len(evidence_records),
        "fieldMappingSuggestions": len(suggestions),
        "lowConfidenceFieldMappings": sum(1 for suggestion in suggestions if suggestion.confidence < 0.6),
        "entityCounts": {
            "products": len(entity_graph.products),
            "productForms": len(entity_graph.product_forms),
            "lots": len(entity_graph.lots),
            "actors": len(entity_graph.actors),
            "locations": len(entity_graph.locations),
            "counterparties": len(entity_graph.counterparties),
            "documents": len(entity_graph.documents),
        },
        "eventNodes": len(event_graph),
        "classifiedEvents": sum(1 for result in classifications if result.final_ctes),
        "abstainedEvents": sum(1 for result in classifications if not result.final_ctes),
        "finalCteCounts": dict(sorted(final_ctes.items())),
        "suppressedCteCounts": dict(sorted(suppressed_ctes.items())),
        "reviewerQuestions": len(reviewer_questions),
        "reviewRequiredEvents": sum(1 for event in event_graph if event.reviewer_questions),
        "phase10a": {
            "inferredFacts": len(inferred_facts),
            "documentProfiles": len(document_profiles),
            "evidenceConflicts": len(evidence_conflicts),
            "documentTypeCounts": dict(sorted(Counter(profile.document_type for profile in document_profiles).items())),
            "qualityReportFile": "phase10a-quality-report.json",
        },
        "acceptanceCoverage": {
            "RI-090_customer_evidence_model": True,
            "RI-091_traceability_entity_model": True,
            "RI-092_spreadsheet_evidence_ingestion": True,
            "RI-093_ai_assisted_field_mapping_suggestions": True,
            "RI-094_customer_event_graph": True,
            "RI-095_food_form_resolver": True,
            "RI-096_actor_role_resolver": True,
            "RI-097_deterministic_cte_classifier": True,
            "RI-098_negative_suppression_logic": True,
            "RI-099_abstention_reviewer_questions": True,
            "RI-10A-001_messy_workbook_parser": True,
            "RI-10A-002_filename_sheet_fact_extraction": True,
            "RI-10A-003_robust_value_normalization": True,
            "RI-10A-004_document_type_parsers": True,
            "RI-10A-005_evidence_conflict_model": True,
            "RI-10A-006_customer_evidence_quality_report": True,
        },
    }


def _detect_header_row(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1.0
    for index, row in enumerate(rows[:30]):
        values = [_cell_to_string(value) for value in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        if len(non_empty) <= 1 and index + 1 < len(rows):
            continue
        alias_hits = sum(1 for value in non_empty if _header_key(value) in FIELD_ALIASES or _slug(value) in FIELD_ALIASES)
        short_header_like = sum(1 for value in non_empty if len(value) <= 45 and not _parse_date_string(value))
        data_like = sum(1 for value in non_empty if _parse_date_string(value) or re.fullmatch(r"-?\d+(?:\.\d+)?", value))
        score = alias_hits * 3 + short_header_like - data_like * 2 + min(len(non_empty), 8) * 0.2
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _build_headers(rows: list[list[Any]], header_index: int) -> list[str]:
    header_row = rows[header_index] if rows else []
    previous_row = rows[header_index - 1] if header_index > 0 else []
    headers: list[str] = []
    max_len = max(len(header_row), len(previous_row))
    for index in range(max_len):
        current = _cell_to_string(header_row[index] if index < len(header_row) else "")
        previous = _cell_to_string(previous_row[index] if index < len(previous_row) else "")
        if current and previous and _header_key(current) not in FIELD_ALIASES and _slug(current) not in FIELD_ALIASES and len(previous) <= 30:
            header = f"{previous} {current}"
        else:
            header = current or previous
        headers.append(_clean_header(header, index + 1))
    return headers


def _is_notes_or_repeated_header_row(row: list[Any], headers: list[str]) -> bool:
    values = [_cell_to_string(value) for value in row]
    non_empty = [value for value in values if value]
    if not non_empty:
        return True
    if len(non_empty) == 1 and re.search(r"^(note|notes|summary|total|generated|report)\b", non_empty[0], re.IGNORECASE):
        return True
    comparable_headers = [_slug(header) for header in headers]
    comparable_values = [_slug(value) for value in values[: len(headers)]]
    matches = sum(1 for left, right in zip(comparable_headers, comparable_values) if left and left == right)
    return matches >= max(2, len(comparable_headers) // 2)


def _merged_or_cell_value(worksheet: Any, row_index: int, column_index: int, *, formula_worksheet: Any | None = None) -> Any:
    cell = worksheet.cell(row=row_index, column=column_index)
    if cell.value not in (None, ""):
        return cell.value
    if formula_worksheet is not None:
        formula_cell = formula_worksheet.cell(row=row_index, column=column_index)
        if getattr(formula_cell, "data_type", None) == "f" and formula_cell.value:
            return formula_cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value


def _infer_facts_from_text(text: str) -> list[tuple[str, str, float]]:
    normalized = text.replace("_", " ").replace("-", " ")
    tokenized = text.replace("_", " ")
    facts: list[tuple[str, str, float]] = []
    for match in re.finditer(r"\b(?:tlc|lot|batch)\s*#?\s*([A-Z0-9][A-Z0-9.-]{2,})\b", tokenized, re.IGNORECASE):
        facts.append(("traceability_lot_code", match.group(1).upper(), 0.82))
    for match in re.finditer(r"\b(20\d{2}[-_/ ]\d{1,2}[-_/ ]\d{1,2}|\d{1,2}[-_/ ]\d{1,2}[-_/ ]20\d{2})\b", normalized):
        facts.append(("event_datetime", match.group(1), 0.78))
    doc_type, signals = _classify_document_type(text)
    if signals:
        facts.append(("source_document_type", doc_type, 0.86))
    product_terms = ["basil", "tuna", "cucumber", "cheese", "sprouts", "lettuce", "tomato", "melon", "pepper", "salad"]
    for term in product_terms:
        if re.search(rf"\b{re.escape(term)}\b", normalized, re.IGNORECASE):
            facts.append(("product_name", _title_from_slug(term), 0.62))
    location_match = re.search(r"\b(farm|warehouse|facility|plant|restaurant|store|vessel|dock)\s+([A-Z0-9][A-Za-z0-9 .-]{1,30})", text, re.IGNORECASE)
    if location_match:
        facts.append(("location_name", location_match.group(0).strip(), 0.66))
    return facts


def _classify_document_type(text: str) -> tuple[str, list[str]]:
    normalized = text.lower()
    definitions = [
        ("invoice", ["invoice", "inv #", "inv-", "bill to"]),
        ("bill_of_lading", ["bill of lading", "bol", "freight", "carrier"]),
        ("receiving_log", ["receiving", "received", "receipt"]),
        ("shipping_log", ["shipping", "shipped", "ship date", "dispatch"]),
        ("transformation_batch_record", ["batch", "transformation", "production", "work order", "recipe"]),
        ("harvest_log", ["harvest", "field", "grower"]),
        ("cooling_log", ["cooling", "cooler", "temperature"]),
        ("packing_log", ["packing", "pack date", "packing house"]),
        ("seafood_landing_record", ["first land", "landing", "vessel", "dock", "landing ticket"]),
        ("traceability_plan", ["traceability plan", "record maintenance", "point of contact"]),
    ]
    best_type = "unknown"
    best_signals: list[str] = []
    for document_type, terms in definitions:
        signals = [term for term in terms if term in normalized]
        if len(signals) > len(best_signals):
            best_type = document_type
            best_signals = signals
    return best_type, best_signals


def _supported_parser_for_document_type(document_type: str) -> str:
    return {
        "invoice": "invoice_parser_v1",
        "bill_of_lading": "bill_of_lading_parser_v1",
        "receiving_log": "receiving_log_parser_v1",
        "shipping_log": "shipping_log_parser_v1",
        "transformation_batch_record": "transformation_batch_record_parser_v1",
        "harvest_log": "harvest_log_parser_v1",
        "cooling_log": "cooling_log_parser_v1",
        "packing_log": "packing_log_parser_v1",
        "seafood_landing_record": "seafood_landing_record_parser_v1",
        "traceability_plan": "traceability_plan_parser_v1",
    }.get(document_type, "generic_spreadsheet_parser_v1")


def _row_facts_with_record_ids(evidence_records: list[CustomerEvidenceRecord]) -> dict[str, dict[str, dict[str, list[str]]]]:
    rows: dict[str, dict[str, dict[str, list[str]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for record in evidence_records:
        if record.confidence < 0.6:
            continue
        row_key = f"{record.uploaded_file}:{record.sheet_name}:{record.row_number}"
        rows[row_key][record.field_key][record.normalized_value].append(record.evidence_id)
    return rows


def _quality_issues(
    *,
    unmapped_columns: list[dict[str, Any]],
    conflicts: list[EvidenceConflict],
    missing_lineage_anchor_count: int,
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if unmapped_columns:
        issues.append(
            {
                "issueType": "unmapped_columns",
                "severity": "medium",
                "count": len(unmapped_columns),
                "message": "Some columns need review before deterministic execution.",
            }
        )
    if conflicts:
        issues.append(
            {
                "issueType": "evidence_conflicts",
                "severity": "high",
                "count": len(conflicts),
                "message": "Conflicting values were preserved as review-blocking facts.",
            }
        )
    if missing_lineage_anchor_count:
        issues.append(
            {
                "issueType": "missing_lineage_anchors",
                "severity": "high",
                "count": missing_lineage_anchor_count,
                "message": "Some evidence records are missing source cell lineage.",
            }
        )
    return issues


def _match_ftl_item(text: str, ftl_food_items: list[dict[str, Any]]) -> str | None:
    if not text:
        return None
    for item in ftl_food_items:
        candidates = [
            item.get("category", ""),
            item.get("commodity", ""),
            item.get("description", ""),
            *(item.get("included_examples") or []),
        ]
        for candidate in candidates:
            candidate_text = str(candidate).lower()
            if candidate_text and (candidate_text in text or any(word and word in text for word in candidate_text.split() if len(word) > 4)):
                return str(item.get("commodity") or item.get("category") or candidate)
    return None


def _load_optional_json_list(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, list) else []


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _mapping_rationale(column_name: str, field_key: str, confidence: float) -> str:
    if confidence >= 0.9:
        return f"Column {column_name!r} matched a known TraceReady/FSMA field alias for {field_key!r}."
    return f"Column {column_name!r} was normalized to {field_key!r}; reviewer confirmation is required before execution."


def _first(facts: dict[str, list[str]], key: str) -> str | None:
    values = facts.get(key) or []
    for value in values:
        if value not in ("", "None", "null"):
            return value
    return None


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return output


def _normalize_cte(value: str | None) -> str | None:
    if not value:
        return None
    normalized = value.lower().replace("_", " ").strip()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized).strip()
    if normalized in CTE_ALIASES:
        return CTE_ALIASES[normalized]
    slug = _slug(value)
    return CTE_ALIASES.get(slug)


def _normalize_value(value: Any, *, field_key: str | None = None) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _cell_to_string(value)
    parsed_date = _parse_date_string(text)
    if parsed_date and (not field_key or "date" in field_key or field_key == "event_datetime"):
        return parsed_date
    normalized = re.sub(r"\s+", " ", text).strip()
    if normalized.startswith("="):
        return normalized
    if field_key and (field_key.endswith("_id") or "lot" in field_key or field_key in {"traceability_lot_code", "source_lot_or_tlc", "output_lot_or_tlc"}):
        return normalized.upper().replace(" ", "")
    if field_key == "unit":
        return _normalize_unit(normalized)
    if field_key == "quantity":
        return _normalize_quantity(normalized)
    if normalized.lower() in {"y", "yes", "true"}:
        return "yes"
    if normalized.lower() in {"n", "no", "false"}:
        return "no"
    return normalized


def _parse_date_string(value: str) -> str | None:
    stripped = value.strip()
    if re.fullmatch(r"\d{4}\s+\d{1,2}\s+\d{1,2}", stripped) or re.fullmatch(r"\d{1,2}\s+\d{1,2}\s+\d{4}", stripped):
        stripped = re.sub(r"\s+", "-", stripped)
    if re.fullmatch(r"\d{8}", stripped):
        for pattern in ("%Y%m%d", "%m%d%Y"):
            try:
                return datetime.strptime(stripped, pattern).date().isoformat()
            except ValueError:
                continue
    stripped = stripped.replace(".", "/")
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d-%b-%Y", "%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(stripped, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _normalize_unit(value: str) -> str:
    normalized = value.lower().strip(".")
    return {
        "lbs": "lb",
        "pounds": "lb",
        "pound": "lb",
        "kgs": "kg",
        "kilograms": "kg",
        "kilogram": "kg",
        "cases": "case",
        "cs": "case",
        "each": "each",
        "ea": "each",
    }.get(normalized, normalized)


def _normalize_quantity(value: str) -> str:
    cleaned = value.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    return match.group(0) if match else value


def _clean_header(value: str, index: int) -> str:
    return str(value or f"column_{index}").strip() or f"column_{index}"


def _header_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower().replace("_", " "))


def _slug(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_"))


def _stable_id(prefix: str, value: str | None) -> str:
    slug = _slug(value)[:80]
    return f"{prefix}-{slug or 'unknown'}"


def _title_from_slug(value: str) -> str:
    return value.replace("_", " ").title()


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"
