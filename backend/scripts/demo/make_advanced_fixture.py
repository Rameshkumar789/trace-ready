"""Generate the ADVANCED adversarial test fixture: Bluegrass Provisions.

A deliberately hostile multi-category export (produce + dairy + seafood + RTE) in a
template that shares NOTHING with the demo workbook or the Sea Eagle export: different
sheet names, different headers, DD/MM/YYYY text dates, merged banner rows, ERP junk rows,
and ~18 planted defects plus false-positive traps the engine must NOT flag.

Outputs (under data/samples/advanced/):
  bluegrass-provisions-export.xlsx   the workbook
  advanced-asn-856.edi               a 5-line X12 856 exercising pre-receipt + door-vs-DB
  advanced-bol.pdf                   a text BOL for the PDF extraction path

Planted defects (the answer key lives in EXPECTED-RESULTS.md next to the outputs):
  W1  4/6 harvest rows missing Field/Block          -> systemic harvest_field_or_container
  W2  Pack Out sheet has NO received-qty column     -> systemic received_quantity_of_rac (6)
  W3  1/3 dock landings missing catch area+trip     -> harvest_range_and_locations (1)
  W4  4/6 receiving rows missing Doc No             -> systemic reference_document_number
  W5  self-receive row (LOC-01 -> LOC-01)           -> self_receive
  W6  cottage lot DL2503050071: 40 in, 55 out       -> mass_balance
  W7  BP2504100033 embeds 10/04, shipped 08/04      -> date_ordering (proves DD/MM parse)
  W8  2505200160 on 2 products, no transformation   -> duplicate_tlc HIGH
  W9  BP2503150041 on 2 SKUs from same batch        -> duplicate_tlc medium (permitted note)
  W10 2411250099 shipped, embeds pre-window date    -> backward_lineage needs_review
  W11 2505120150 + 2506010161 shipped, in-window,
      no origin                                     -> backward_lineage GAP (high)
  W12 2502140012 packed 14/02, never moved          -> forward_linkage unmoved
  W13 2 production rows missing Primary Input Lot   -> 2 individual source-TLC findings
  W14 plan: point-of-contact blank, update/retention
      row absent                                    -> traceability_plan (2 missing)
  W15 wrap GTIN check digit deliberately wrong      -> gs1_identifier (+ retailer overlays)
  W16 3 products declared "General grocery" but on
      FTL (fresh-cut kit, smoked salmon, mahi-mahi) -> 3 ftl_declared_mismatch
  W17 frozen brie declared as FTL soft cheese       -> frozen-cheese guard => definite_off
  W18 IMS Grade A cottage cheese                    -> cottage guard => never definite_on
  W19 "Asst. Deli Cups" shipped, not in catalog     -> scope_or_exemption_uncertainty

False-positive traps (must NOT fire):
  T1  transformation outputs carry Primary Input Lot -> NO transformation_linkage finding
  T2  lot 2502140011 partially consumed              -> NO mass-balance false alarm
  T3  BP2505220055/56 + DL2506150090 originate within
      30 days of window end, unshipped               -> NOT in forward-linkage unmoved list
  T4  cooling sheet fully compliant                  -> zero cooling findings
  T5  frozen mahi/smoked salmon                      -> frozen guard must NOT clear seafood
  T6  DD/MM dates                                    -> window must be Feb..Jun 2025, not
                                                        May..Dec (US-locale misparse)

EDI (advanced-asn-856.edi, Dairyland -> Bluegrass):
  L1 cottage DL2503050071 qty 40  -> pre-receipt accept; audit diff: supplier sends REF/
                                     phone/email the ERP dropped (dropped-fields finding)
  L2 brie    DL2503050072 qty 70  -> pre-receipt accept; audit diff: qty conflict vs 60 (gap)
  L3 cottage DL2503990099 qty 20  -> pre-receipt accept; audit diff: lot unknown to system
  L4 peanut butter, NO lot        -> pre-receipt HOLD (missing TLC)
  L5 cottage DL2506150090 qty 35  -> pre-receipt accept; matches R-106 exactly

Run from backend/:  python scripts/demo/make_advanced_fixture.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parents[3]
OUT = REPO / "data/samples/advanced"


# --------------------------------------------------------------------------- GTIN helpers
def gs1_check_digit(digits: str) -> str:
    total = 0
    for i, ch in enumerate(reversed(digits)):
        total += int(ch) * (3 if i % 2 == 0 else 1)
    return str((10 - total % 10) % 10)


def gtin14(seq: int) -> str:
    body = f"1086000452{seq:03d}"  # 13 digits
    return body + gs1_check_digit(body)


ITEMS = {
    # key: (gtin, description, declared_category, storage, unit)
    "romaine": (gtin14(1), "Romaine Hearts 3ct Fresh", "Leafy greens (fresh)", "Refrigerated", "case"),
    "salad_kit": (gtin14(2), "Fresh-Cut Caesar Salad Kit with Dressing", "General grocery", "Refrigerated", "case"),
    "smoked_salmon": (gtin14(3), "Cold Smoked Atlantic Salmon Slices 8oz", "General grocery", "Refrigerated", "case"),
    "frozen_brie": (gtin14(4), "Frozen Brie Wheel 1kg (soft ripened, pasteurized)", "Cheese (made from pasteurized milk), soft ripened or semi-soft", "Frozen", "case"),
    "cottage": (gtin14(5), "Cottage Cheese Grade A 16oz (IMS listed)", "Cheese (made from pasteurized milk), fresh soft or soft unripened", "Refrigerated", "case"),
    "peanut_butter": (gtin14(6), "Creamy Peanut Butter 40lb Pail", "Nut butters", "Ambient", "case"),
    "oregano": (gtin14(7), "Dried Oregano 1lb", "General grocery", "Ambient", "case"),
    "tomatoes": (gtin14(8), "Vine Tomatoes 25lb", "Tomatoes (fresh)", "Refrigerated", "case"),
    "mahi": (gtin14(9), "Mahi-Mahi Fillets IQF Frozen", "General grocery", "Frozen", "lb"),
    "canned_tuna": (gtin14(10), "Canned Tuna in Water (shelf-stable)", "General grocery", "Ambient", "case"),
    # W15: wrap GTIN check digit deliberately broken (+1 mod 10)
    "wrap": (gtin14(11)[:-1] + str((int(gtin14(11)[-1]) + 1) % 10), "Chicken Caesar Wrap RTE", "General grocery", "Refrigerated", "case"),
}


def _banner(ws, title: str, width: int) -> int:
    """Merged banner + ERP junk row + blank row; returns the header row index."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Generated by BluePrint ERP v4.2 • 30/06/2025 • CONFIDENTIAL")
    return 4


def _rows(ws, header_row: int, headers: list[str], rows: list[list]) -> None:
    for col, name in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=name).font = Font(bold=True)
    for r, row in enumerate(rows, start=header_row + 1):
        for c, value in enumerate(row, start=1):
            if value is not None and value != "":
                ws.cell(row=r, column=c, value=value)


def build_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Company Register")
    _rows(ws, _banner(ws, "BLUEGRASS PROVISIONS LLC — TRACEABILITY EXPORT", 4), ["Company Ref", "Legal Name", "Role", "City"], [
        ["BPL-001", "Bluegrass Provisions LLC", "Operator (processor/distributor)", "Louisville, KY"],
        ["GAF-002", "GreenAcre Farms", "Supplier - grower", "Shelbyville, KY"],
        ["LKF-003", "Lakeshore Fisheries Co-op", "Supplier - first land-based receiver", "Pascagoula, MS"],
        ["DLC-004", "Dairyland Creamery", "Supplier - dairy", "Madison, WI"],
        ["NWI-005", "NutWorks Ingredients", "Supplier - ambient", "Albany, GA"],
        ["MGR-006", "Metro Grocers Inc", "Customer - retail", "Cincinnati, OH"],
        ["SRG-007", "Sunbelt Restaurants Group", "Customer - foodservice", "Nashville, TN"],
    ])

    ws = wb.create_sheet("Facility Register")
    _rows(ws, _banner(ws, "FACILITIES", 6), ["Location Ref", "Facility Name", "Street", "City", "State", "Operated By"], [
        ["LOC-01", "Louisville Distribution Center", "400 River Rd", "Louisville", "KY", "Bluegrass Provisions LLC"],
        ["LOC-02", "Riverside Processing Plant", "88 Mill St", "Louisville", "KY", "Bluegrass Provisions LLC"],
        ["LOC-03", "GreenAcre Home Farm", "1200 County Rd 9", "Shelbyville", "KY", "GreenAcre Farms"],
        ["LOC-04", "Lakeshore Dock 7", "Pier 7, Harbor Dr", "Pascagoula", "MS", "Lakeshore Fisheries Co-op"],
    ])

    ws = wb.create_sheet("Item Catalog")
    _rows(ws, _banner(ws, "ITEM CATALOG", 5), ["Item Code", "Item Description", "Declared FTL Category", "Storage", "Unit"], [
        [g, d, c, s, u] for (g, d, c, s, u) in ITEMS.values()
    ])

    romaine, tomatoes = ITEMS["romaine"], ITEMS["tomatoes"]
    ws = wb.create_sheet("Harvest Log")
    _rows(ws, _banner(ws, "HARVEST LOG — GREENACRE FARMS", 13),
          ["Harvest Ref", "Crop Item", "Crop Description", "Qty Harvested", "UOM", "Date of Harvest", "Farm", "Field or Block", "Crew Lead", "Crew Phone", "Shipped To", "Doc Type", "Doc Ref"], [
        ["H-01", romaine[0], "Romaine Hearts 3ct Fresh", 110, "case", "05/02/2025", "GreenAcre Home Farm", "Block A-3", "M. Vasquez", "502-555-0161", "Riverside Processing Plant", "Harvest ticket", "HT-3301"],
        ["H-02", tomatoes[0], "Vine Tomatoes 25lb", 90, "case", "14/02/2025", "GreenAcre Home Farm", "Greenhouse 2", "M. Vasquez", "502-555-0161", "Riverside Processing Plant", "Harvest ticket", "HT-3302"],
        # W1: 4 rows below missing Field or Block -> systemic gap
        ["H-03", romaine[0], "Romaine Hearts 3ct Fresh", 55, "case", "21/02/2025", "GreenAcre Home Farm", "", "J. Osei", "502-555-0162", "Riverside Processing Plant", "Harvest ticket", "HT-3303"],
        ["H-04", tomatoes[0], "Vine Tomatoes 25lb", 65, "case", "01/03/2025", "GreenAcre Home Farm", "", "J. Osei", "502-555-0162", "Riverside Processing Plant", "Harvest ticket", "HT-3304"],
        ["H-05", romaine[0], "Romaine Hearts 3ct Fresh", 45, "case", "08/03/2025", "GreenAcre Home Farm", "", "M. Vasquez", "502-555-0161", "Riverside Processing Plant", "Harvest ticket", "HT-3305"],
        ["H-06", tomatoes[0], "Vine Tomatoes 25lb", 35, "case", "08/03/2025", "GreenAcre Home Farm", "", "M. Vasquez", "502-555-0161", "Riverside Processing Plant", "Harvest ticket", "HT-3306"],
    ])

    # T4: cooling fully compliant (native datetime cells exercise the Excel-date path)
    ws = wb.create_sheet("Cooling Register")
    _rows(ws, _banner(ws, "COOLING REGISTER", 11),
          ["Cooling Ref", "Item", "Description", "Qty", "UOM", "Cooling Date", "Cooling Site", "Farm of Origin", "Next Recipient", "Doc Type", "Doc Ref"], [
        ["C-01", romaine[0], "Romaine Hearts 3ct Fresh", 110, "case", datetime(2025, 2, 5), "GreenAcre Home Farm", "GreenAcre Home Farm", "Riverside Processing Plant", "Cooling log", "CL-901"],
        ["C-02", romaine[0], "Romaine Hearts 3ct Fresh", 55, "case", datetime(2025, 2, 21), "GreenAcre Home Farm", "GreenAcre Home Farm", "Riverside Processing Plant", "Cooling log", "CL-902"],
        ["C-03", romaine[0], "Romaine Hearts 3ct Fresh", 45, "case", datetime(2025, 3, 8), "GreenAcre Home Farm", "GreenAcre Home Farm", "Riverside Processing Plant", "Cooling log", "CL-903"],
    ])

    # W2: no received-qty column at all -> systemic received_quantity_of_rac
    ws = wb.create_sheet("Pack Out")
    _rows(ws, _banner(ws, "PACK OUT — INITIAL PACKING", 16),
          ["Pack Ref", "Item", "Description", "Lot Assigned", "Qty Packed", "UOM", "Pack Date", "Packing Site", "Farm", "Field or Block", "Harvest Date", "Received Date", "Harvester Name", "Harvester Phone", "Doc Type", "Doc Ref"], [
        ["P-01", romaine[0], "Romaine Hearts 3ct Fresh", "2502140011", 100, "case", "14/02/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Block A-3", "05/02/2025", "06/02/2025", "M. Vasquez", "502-555-0161", "Pack ticket", "PK-501"],
        ["P-02", tomatoes[0], "Vine Tomatoes 25lb", "2502140012", 80, "case", "14/02/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Greenhouse 2", "14/02/2025", "14/02/2025", "M. Vasquez", "502-555-0161", "Pack ticket", "PK-502"],
        ["P-03", romaine[0], "Romaine Hearts 3ct Fresh", "2502210013", 50, "case", "21/02/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Block B-1", "21/02/2025", "21/02/2025", "J. Osei", "502-555-0162", "Pack ticket", "PK-503"],
        ["P-04", tomatoes[0], "Vine Tomatoes 25lb", "2503010014", 60, "case", "01/03/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Greenhouse 2", "01/03/2025", "01/03/2025", "J. Osei", "502-555-0162", "Pack ticket", "PK-504"],
        ["P-05", romaine[0], "Romaine Hearts 3ct Fresh", "2503080015", 40, "case", "08/03/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Block A-3", "08/03/2025", "08/03/2025", "M. Vasquez", "502-555-0161", "Pack ticket", "PK-505"],
        ["P-06", tomatoes[0], "Vine Tomatoes 25lb", "2503080016", 30, "case", "08/03/2025", "GreenAcre Home Farm", "GreenAcre Home Farm", "Greenhouse 2", "08/03/2025", "08/03/2025", "M. Vasquez", "502-555-0161", "Pack ticket", "PK-506"],
    ])

    mahi = ITEMS["mahi"]
    ws = wb.create_sheet("Dock Landings")
    _rows(ws, _banner(ws, "DOCK LANDINGS — FIRST LAND-BASED RECEIVER", 12),
          ["Landing Ref", "Species", "Market Name", "Lot", "Qty", "UOM", "Landed Date", "Catch Area (FAO)", "Vessel Trip Dates", "Receiver", "Doc Type", "Doc Ref"], [
        ["LD-01", mahi[0], "Mahi-Mahi (dolphinfish)", "LK2502250021", 600, "lb", "25/02/2025", "FAO 31 - Gulf of Mexico", "18/02/2025-24/02/2025", "Lakeshore Dock 7", "Landing ticket", "LT-771"],
        ["LD-02", mahi[0], "Mahi-Mahi (dolphinfish)", "LK2502250022", 550, "lb", "25/02/2025", "FAO 31 - Gulf of Mexico", "18/02/2025-24/02/2025", "Lakeshore Dock 7", "Landing ticket", "LT-772"],
        # W3: catch area AND trip dates blank
        ["LD-03", mahi[0], "Mahi-Mahi (dolphinfish)", "LK2503200023", 400, "lb", "20/03/2025", "", "", "Lakeshore Dock 7", "Landing ticket", "LT-773"],
    ])

    cottage, brie, pb = ITEMS["cottage"], ITEMS["frozen_brie"], ITEMS["peanut_butter"]
    ws = wb.create_sheet("Goods In")
    _rows(ws, _banner(ws, "GOODS IN — RECEIVING", 12),
          ["Receipt Ref", "Item", "Description", "Lot Received", "Qty", "UOM", "Received On", "From Partner", "From Site", "To Site", "Doc Type", "Doc No"], [
        # W4: 4 rows missing Doc No
        ["R-101", cottage[0], "Cottage Cheese Grade A 16oz (IMS listed)", "DL2503050071", 40, "case", "05/03/2025", "Dairyland Creamery", "Dairyland Plant 2", "Louisville Distribution Center", "Invoice", ""],
        ["R-102", brie[0], "Frozen Brie Wheel 1kg", "DL2503050072", 60, "case", "05/03/2025", "Dairyland Creamery", "Dairyland Plant 2", "Louisville Distribution Center", "Invoice", ""],
        ["R-103", pb[0], "Creamy Peanut Butter 40lb Pail", "NW2502200005", 50, "case", "20/02/2025", "NutWorks Ingredients", "NutWorks Albany", "Louisville Distribution Center", "Invoice", ""],
        ["R-104", tomatoes[0], "Vine Tomatoes 25lb", "2505200160", 80, "case", "20/05/2025", "GreenAcre Farms", "GreenAcre Home Farm", "Louisville Distribution Center", "Invoice", ""],
        # W5: self-receive
        ["R-105", romaine[0], "Romaine Hearts 3ct Fresh", "2504010030", 25, "case", "01/04/2025", "Bluegrass Provisions LLC", "Louisville Distribution Center", "Louisville Distribution Center", "Transfer", "TRF-2211"],
        # T3: window-end carve-out lot (received 15/06, never shipped, must NOT read unmoved)
        ["R-106", cottage[0], "Cottage Cheese Grade A 16oz (IMS listed)", "DL2506150090", 35, "case", "15/06/2025", "Dairyland Creamery", "Dairyland Plant 2", "Louisville Distribution Center", "Invoice", "INV-8890"],
    ])

    salad, wrap = ITEMS["salad_kit"], ITEMS["wrap"]
    ws = wb.create_sheet("Batch Inputs")
    _rows(ws, _banner(ws, "BATCH INPUTS — INGREDIENTS CONSUMED", 7),
          ["Batch No", "Input Item", "Input Description", "Lot Consumed", "Qty Used", "UOM", "Usage Date"], [
        ["B-0315", romaine[0], "Romaine Hearts 3ct Fresh", "2502140011", 30, "case", "15/03/2025"],
        ["B-0315", "EXT-DRS-889", "Caesar Dressing Bulk (non-FTL)", "EXTDRS889", 5, "case", "15/03/2025"],
        ["B-0410", romaine[0], "Romaine Hearts 3ct Fresh", "2502140011", 10, "case", "10/04/2025"],
        ["B-0410", "EXT-DRS-890", "Caesar Dressing Bulk (non-FTL)", "EXTDRS890", 2, "case", "10/04/2025"],
    ])

    ws = wb.create_sheet("Production Batches")
    # One row per output SKU with a unique line ref: rows sharing an event-id value are
    # merged into one multi-line event by design, which is not what a per-SKU export means.
    _rows(ws, _banner(ws, "PRODUCTION BATCHES — OUTPUTS", 12),
          ["Output Line Ref", "Batch No", "Output Item", "Output Description", "New Lot", "Primary Input Lot", "Qty Produced", "UOM", "Production Date", "Production Site", "Doc Type", "Doc Ref"], [
        # T1: source lots present -> no transformation_linkage finding
        ["B-0315-A", "B-0315", salad[0], "Fresh-Cut Caesar Salad Kit with Dressing", "BP2503150041", "2502140011", 120, "case", "15/03/2025", "Riverside Processing Plant", "Batch record", "BR-0315"],
        # W9: same batch, same lot, second SKU (permitted multi-SKU note)
        ["B-0315-B", "B-0315", wrap[0], "Chicken Caesar Wrap RTE", "BP2503150041", "2502140011", 80, "case", "15/03/2025", "Riverside Processing Plant", "Batch record", "BR-0315"],
        ["B-0410-A", "B-0410", salad[0], "Fresh-Cut Caesar Salad Kit with Dressing", "BP2504100033", "2502140011", 45, "case", "10/04/2025", "Riverside Processing Plant", "Batch record", "BR-0410"],
        # W13: two rows missing the input lot (individual findings, below rollup threshold)
        ["B-0522-A", "B-0522", wrap[0], "Chicken Caesar Wrap RTE", "BP2505220055", "", 200, "case", "22/05/2025", "Riverside Processing Plant", "Batch record", "BR-0522"],
        ["B-0529-A", "B-0529", wrap[0], "Chicken Caesar Wrap RTE", "BP2505290056", "", 150, "case", "29/05/2025", "Riverside Processing Plant", "Batch record", "BR-0529"],
    ])

    smoked = ITEMS["smoked_salmon"]
    ws = wb.create_sheet("Dispatch Log")
    _rows(ws, _banner(ws, "DISPATCH LOG — OUTBOUND", 12),
          ["Ship Ref", "Item", "Description", "Lot Shipped", "Qty", "UOM", "Ship Date", "Ship From", "Customer", "Ship To City", "Doc Type", "Doc No"], [
        ["S-201", romaine[0], "Romaine Hearts 3ct Fresh", "2502140011", 60, "case", "18/02/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7001"],
        ["S-202", romaine[0], "Romaine Hearts 3ct Fresh", "2502210013", 50, "case", "24/02/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7002"],
        ["S-203", tomatoes[0], "Vine Tomatoes 25lb", "2503010014", 60, "case", "04/03/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7003"],
        ["S-204", romaine[0], "Romaine Hearts 3ct Fresh", "2503080015", 40, "case", "12/03/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7004"],
        ["S-205", tomatoes[0], "Vine Tomatoes 25lb", "2503080016", 30, "case", "12/03/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7005"],
        ["S-206", salad[0], "Fresh-Cut Caesar Salad Kit with Dressing", "BP2503150041", 70, "case", "17/03/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7006"],
        ["S-207", wrap[0], "Chicken Caesar Wrap RTE", "BP2503150041", 40, "case", "18/03/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7007"],
        # W7: lot embeds 2025-04-10, shipped 08/04 (DD/MM) -> date_ordering; misparse kills this trap
        ["S-208", salad[0], "Fresh-Cut Caesar Salad Kit with Dressing", "BP2504100033", 20, "case", "08/04/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7008"],
        # W6: 30 + 25 = 55 shipped vs 40 received
        ["S-209", cottage[0], "Cottage Cheese Grade A 16oz (IMS listed)", "DL2503050071", 30, "case", "10/03/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7009"],
        ["S-210", cottage[0], "Cottage Cheese Grade A 16oz (IMS listed)", "DL2503050071", 25, "case", "21/03/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7010"],
        ["S-211", brie[0], "Frozen Brie Wheel 1kg", "DL2503050072", 30, "case", "15/03/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7011"],
        ["S-212", pb[0], "Creamy Peanut Butter 40lb Pail", "NW2502200005", 20, "case", "25/02/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7012"],
        ["S-213", mahi[0], "Mahi-Mahi Fillets IQF Frozen", "LK2502250021", 600, "lb", "28/02/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7013"],
        ["S-214", mahi[0], "Mahi-Mahi Fillets IQF Frozen", "LK2502250022", 550, "lb", "01/03/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7014"],
        ["S-215", tomatoes[0], "Vine Tomatoes 25lb", "2505200160", 40, "case", "22/05/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7015"],
        # W8: same lot, different product, no transformation
        ["S-216", romaine[0], "Romaine Hearts 3ct Fresh", "2505200160", 30, "case", "23/05/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7016"],
        # W10: pre-window orphan (embeds 2024-11-25)
        ["S-217", smoked[0], "Cold Smoked Atlantic Salmon Slices 8oz", "2411250099", 15, "case", "20/02/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7017"],
        # W11: in-window orphans -> true backward-lineage gaps
        ["S-218", mahi[0], "Mahi-Mahi Fillets IQF Frozen", "2505120150", 200, "lb", "14/05/2025", "Louisville Distribution Center", "Sunbelt Restaurants Group", "Nashville, TN", "BOL", "BG-7018"],
        # W19: product not in catalog, ambiguous description (also W11's second orphan lot)
        ["S-219", "", "Asst. Deli Cups 12ct", "2506010161", 30, "case", "01/06/2025", "Louisville Distribution Center", "Metro Grocers Inc", "Cincinnati, OH", "BOL", "BG-7019"],
    ])

    # W14: point-of-contact answered blank; update/retention row absent entirely
    ws = wb.create_sheet("Trace Plan")
    _rows(ws, _banner(ws, "TRACEABILITY PLAN — 21 CFR 1.1315", 2), ["Plan Component", "Details"], [
        ["How records are maintained (format and location)", "Records live in BluePrint ERP; signed PDFs archived on SharePoint /Compliance; retained 2 years."],
        ["How we identify Food Traceability List foods", "Item master carries a Declared FTL Category reviewed quarterly against the FDA FTL."],
        ["How traceability lot codes are assigned", "Supplier lots kept as-is; production lots = BP + YYMMDD + batch sequence."],
        ["Point of contact for traceability questions", ""],
        ["Farm map of growing areas", "GreenAcre field maps on file: GA-MAP-01 (Block A), GA-MAP-02 (Greenhouses)."],
    ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"wrote {path}")


def build_asn(path: Path) -> None:
    cottage, brie, pb = ITEMS["cottage"], ITEMS["frozen_brie"], ITEMS["peanut_butter"]
    segments = [
        "ISA*00*          *00*          *ZZ*DAIRYLANDCRMY  *ZZ*BLUEGRASSPROV  *250305*0615*U*00401*000000207*0*P*>",
        "GS*SH*DAIRYLANDCRMY*BLUEGRASSPROV*20250305*0615*207*X*004010",
        "ST*856*0001",
        "BSN*00*ASN-55107*20250305*061500",
        "DTM*011*20250305",
        "N1*SF*Dairyland Creamery*92*DLC-004",
        "PER*IC*Shipping Office*TE*608-555-0177*EM*ship@dairylandcreamery.example",
        "N1*ST*Bluegrass Provisions LLC*92*BPL-001",
        "REF*BM*BOL-55107",
        "HL*1**S",
        "TD1*CTN*205",
        "HL*2*1*O",
        "HL*3*2*I",
        f"LIN**UP*{cottage[0]}*LT*DL2503050071",
        "SN1**40*CA",
        "PID*F****Cottage Cheese Grade A 16oz (IMS listed)",
        "HL*4*2*I",
        f"LIN**UP*{brie[0]}*LT*DL2503050072",
        "SN1**70*CA",
        "PID*F****Frozen Brie Wheel 1kg",
        "HL*5*2*I",
        f"LIN**UP*{cottage[0]}*LT*DL2503990099",
        "SN1**20*CA",
        "PID*F****Cottage Cheese Grade A 16oz (IMS listed)",
        "HL*6*2*I",
        f"LIN**UP*{pb[0]}",
        "SN1**25*CA",
        "PID*F****Creamy Peanut Butter 40lb Pail",
        "HL*7*2*I",
        f"LIN**UP*{cottage[0]}*LT*DL2506150090",
        "SN1**35*CA",
        "PID*F****Cottage Cheese Grade A 16oz (IMS listed)",
        "CTT*5",
        "SE*32*0001",
        "GE*1*207",
        "IEA*1*000000207",
    ]
    path.write_text("~\n".join(segments) + "~\n", encoding="utf-8")
    print(f"wrote {path}")


def build_bol(path: Path) -> None:
    import fitz  # PyMuPDF

    cottage, brie = ITEMS["cottage"], ITEMS["frozen_brie"]
    doc = fitz.open()
    page = doc.new_page()
    y = 60
    lines = [
        ("STRAIGHT BILL OF LADING — NOT NEGOTIABLE", 14, True),
        ("BOL Number: BOL-55107        Date: 03/05/2025", 10, False),
        ("Carrier: Cold Chain Express   PRO: CCX-99120", 10, False),
        ("", 10, False),
        ("SHIP FROM: Dairyland Creamery, Plant 2, 500 Creamery Way, Madison, WI 53703", 10, False),
        ("Phone: 608-555-0177    Email: ship@dairylandcreamery.example", 10, False),
        ("SHIP TO: Bluegrass Provisions LLC, Louisville Distribution Center, 400 River Rd, Louisville, KY 40202", 10, False),
        ("", 10, False),
        ("LINE  PRODUCT                                          GTIN              LOT NO         QTY   UOM", 9, True),
        (f"1     Cottage Cheese Grade A 16oz (IMS listed)        {cottage[0]}    DL2503050071   40    CASE", 9, False),
        (f"2     Frozen Brie Wheel 1kg                           {brie[0]}    DL2503050072   70    CASE", 9, False),
        (f"3     Cottage Cheese Grade A 16oz (IMS listed)        {cottage[0]}    DL2503990099   20    CASE", 9, False),
        ("", 10, False),
        ("Special instructions: keep 34-38F. Lot codes must appear on receiving record per FSMA 204.", 9, False),
        ("Shipper signature: R. Halvorsen      Driver: T. Boone      Trailer: 5521", 9, False),
    ]
    for text, size, bold in lines:
        if text:
            page.insert_text((50, y), text, fontsize=size, fontname="courier-bold" if bold else "courier")
        y += size + 8
    doc.save(path)
    print(f"wrote {path}")


if __name__ == "__main__":
    build_workbook(OUT / "bluegrass-provisions-export.xlsx")
    build_asn(OUT / "advanced-asn-856.edi")
    build_bol(OUT / "advanced-bol.pdf")
