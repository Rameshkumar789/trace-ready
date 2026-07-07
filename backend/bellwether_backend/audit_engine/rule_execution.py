from __future__ import annotations

import json
from collections import Counter, OrderedDict, defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Any

from pydantic import BaseModel, ConfigDict, Field

from bellwether_backend.audit_engine.cte_classification import (
    MultiSignalCteResult,
    build_phase10c_cte_hardening,
)
from bellwether_backend.audit_engine.customer_evidence import (
    CustomerEventNode,
    CustomerEvidenceRecord,
    Phase10CustomerEvidencePackage,
    build_phase10_customer_evidence,
)


GENERATED_AT = "2026-06-16T00:00:00Z"


class StrictRuleExecutionModel(BaseModel):
    model_config = ConfigDict(extra="forbid", validate_assignment=True, str_strip_whitespace=True)


class EventObligationMapping(StrictRuleExecutionModel):
    mapping_id: str
    event_id: str
    cte: str
    approved_obligation_id: str
    obligation_action: str
    required_output: str | None = None
    citation: dict[str, Any]
    rule_package_id: str
    rule_package_version: int


class KdeCompletenessCheck(StrictRuleExecutionModel):
    check_id: str
    event_id: str
    cte: str
    field_key: str
    status: str
    expected_reason: str
    evidence_ids: list[str] = Field(default_factory=list)
    observed_values: list[str] = Field(default_factory=list)
    approved_obligation_id: str


class TlcLineageCheck(StrictRuleExecutionModel):
    check_id: str
    event_id: str
    cte: str
    status: str
    lot_or_tlc: str | None = None
    source_lot_or_tlc: str | None = None
    output_lot_or_tlc: str | None = None
    evidence_ids: list[str] = Field(default_factory=list)
    approved_obligation_id: str | None = None
    reason: str


class TraceabilityPlanCheck(StrictRuleExecutionModel):
    check_id: str
    component: str
    status: str
    evidence_ids: list[str] = Field(default_factory=list)
    approved_obligation_id: str
    reason: str


class ScopeExemptionCheck(StrictRuleExecutionModel):
    check_id: str
    event_id: str | None = None
    check_type: str
    status: str
    reason: str
    evidence_ids: list[str] = Field(default_factory=list)


class RecordsReadinessCheck(StrictRuleExecutionModel):
    check_id: str
    check_type: str
    status: str
    reason: str
    evidence_ids: list[str] = Field(default_factory=list)
    approved_obligation_id: str


class SortableExportReadinessCheck(StrictRuleExecutionModel):
    check_id: str
    event_id: str
    cte: str
    status: str
    missing_fields: list[str] = Field(default_factory=list)
    populated_fields: list[str] = Field(default_factory=list)
    approved_obligation_id: str


class AuditFinding(StrictRuleExecutionModel):
    finding_id: str
    event_id: str | None = None
    cte: str | None = None
    severity: str
    status: str
    finding_type: str
    message: str
    approved_obligation_id: str
    source_citation: dict[str, Any]
    customer_evidence_ids: list[str] = Field(default_factory=list)
    confidence: float = Field(ge=0, le=1)
    reviewer_status: str
    # Specific gaps rolled up into this single record-level finding (human-readable),
    # so one root cause is reported once instead of as several near-duplicate findings.
    sub_issues: list[str] = Field(default_factory=list)
    affected_fields: list[str] = Field(default_factory=list)
    # Whether the requirement comes from the FDA rule or from a specific customer's supplier
    # instructions (retailer overlay), e.g. Walmart's GS1 mandate.
    requirement_source: str = "fda_rule"


class ExceptionQueueItem(StrictRuleExecutionModel):
    exception_id: str
    finding_id: str | None = None
    event_id: str | None = None
    queue_type: str
    priority: str
    title: str
    details: str
    evidence_ids: list[str] = Field(default_factory=list)
    assigned_role: str
    status: str = "open"


class FdaStyleExportPackage(StrictRuleExecutionModel):
    package_id: str
    generated_at: str
    rule_package_id: str
    rule_package_version: int
    status: str
    workbook_file: str
    tabs: dict[str, list[dict[str, Any]]]
    blockers: list[dict[str, Any]]
    citations: list[dict[str, Any]]


class Phase11RuleExecutionPackage(StrictRuleExecutionModel):
    generated_at: str
    summary: dict[str, Any]
    obligation_mappings: list[EventObligationMapping]
    kde_checks: list[KdeCompletenessCheck]
    tlc_checks: list[TlcLineageCheck]
    traceability_plan_checks: list[TraceabilityPlanCheck]
    scope_exemption_checks: list[ScopeExemptionCheck]
    records_readiness_checks: list[RecordsReadinessCheck]
    sortable_export_checks: list[SortableExportReadinessCheck]
    audit_findings: list[AuditFinding]
    exception_queue: list[ExceptionQueueItem]
    export_package: FdaStyleExportPackage
    # Additive (defaulted) so existing consumers keep working unchanged.
    lot_integrity_checks: list[Any] = Field(default_factory=list)
    gs1_checks: list[Any] = Field(default_factory=list)
    ftl_tier_results: dict[str, Any] = Field(default_factory=dict)
    partner_scorecard: dict[str, Any] = Field(default_factory=dict)
    scoping_report: dict[str, Any] = Field(default_factory=dict)
    mapping_plan: dict[str, Any] | None = None


BUNDLED_RULES_DIR = Path(__file__).resolve().parent / "bundled_rules"
KDE_CHECK_CONTRACTS_PATH = BUNDLED_RULES_DIR / "kde-check-contracts.json"


@lru_cache(maxsize=8)
def _load_kde_check_contracts(path: Path | None = None) -> dict[str, dict[str, Any]]:
    """The full FSMA 204 KDE list per CTE, each mapped to the customer field(s) the parser
    produces. Source of truth is the approved kde_check_contracts cards in Supabase, written
    to a run-scoped file and passed in; the bundled JSON is only an offline/dev fallback. Not
    a hardcoded code list, so coverage stays complete and in sync with the regulation."""
    data = json.loads((path or KDE_CHECK_CONTRACTS_PATH).read_text(encoding="utf-8"))
    return data["cte_contracts"]


EXEMPTION_RULES_PATH = BUNDLED_RULES_DIR / "exemption-rules.json"


@lru_cache(maxsize=8)
def _load_exemption_rules(path: Path | None = None) -> list[dict[str, Any]]:
    """FSMA 204 exemption rules (21 CFR 1.1305). Source of truth is the approved
    exemption_rules cards in Supabase, written to a run-scoped file and passed in; the bundled
    JSON is only an offline/dev fallback. Reviewable data, not code."""
    data = json.loads((path or EXEMPTION_RULES_PATH).read_text(encoding="utf-8"))
    return data["exemptions"]


PLAN_COMPONENTS_PATH = BUNDLED_RULES_DIR / "traceability-plan-components.json"


@lru_cache(maxsize=8)
def _load_plan_components(path: Path | None = None) -> list[dict[str, Any]]:
    """The traceability-plan components a covered entity must document (21 CFR 1.1315). Source
    of truth is the approved traceability_plan_components cards in Supabase, written to a
    run-scoped file and passed in; the bundled JSON is only an offline/dev fallback."""
    data = json.loads((path or PLAN_COMPONENTS_PATH).read_text(encoding="utf-8"))
    return data["components"]

CTE_SECTION_REFS = {
    "harvesting": "21 CFR 1.1325",
    "cooling": "21 CFR 1.1325",
    "initial_packing": "21 CFR 1.1330",
    "first_land_based_receiving": "21 CFR 1.1335",
    "shipping": "21 CFR 1.1340",
    "receiving": "21 CFR 1.1345",
    "transformation": "21 CFR 1.1350",
}

# Grower/harvester-style CTEs. A traceability-plan "farm map" is only required of these
# operations (21 CFR 1.1315(a)(4)); for a distributor/processor it is not applicable.
FARM_OPERATION_CTES = {"harvesting", "cooling", "initial_packing"}

# TLC fields are the "root cause" lever: a missing TLC tends to trigger both a
# KDE-completeness gap and a TLC-lineage gap on the same record. We collapse those
# into one record-level finding instead of counting the same problem multiple times.
TLC_FIELD_KEYS = {"traceability_lot_code", "output_lot_or_tlc", "source_lot_or_tlc"}
HIGH_SEVERITY_FIELD_KEYS = {"traceability_lot_code", "output_lot_or_tlc"}

# Partner-facing names so findings read in plain English, not internal field keys.
FRIENDLY_CTE_LABELS = {
    "shipping": "Shipping record",
    "receiving": "Receiving record",
    "transformation": "Transformation record",
    "first_land_based_receiving": "First land-based receiving record",
    "initial_packing": "Initial packing record",
    "harvesting": "Harvest record",
    "cooling": "Cooling record",
    "traceability_plan": "Traceability plan",
}

FRIENDLY_FIELD_LABELS = {
    "traceability_lot_code": "Traceability Lot Code (TLC)",
    "output_lot_or_tlc": "new Traceability Lot Code for the transformed product",
    "source_lot_or_tlc": "source Traceability Lot Code for the incoming product",
    "product_name": "product description",
    "event_datetime": "event date",
    "reference_record_type": "reference document type",
    "reference_record_no": "reference document number",
}

FRIENDLY_PLAN_COMPONENT_LABELS = {
    "ftl_food_identification": "procedure to identify Food Traceability List foods",
    "tlc_assignment_procedure": "Traceability Lot Code assignment procedure",
    "point_of_contact": "traceability point of contact",
    "farm_map": "farm map",
    "plan_update_and_retention": "plan update and record-retention procedure",
    "record_maintenance_procedure": "record maintenance procedure",
}


def _cte_label(cte: str | None) -> str:
    if not cte:
        return "Record"
    return FRIENDLY_CTE_LABELS.get(cte, cte.replace("_", " ").strip().title())


def _field_label(field_key: str) -> str:
    return FRIENDLY_FIELD_LABELS.get(field_key, field_key.replace("_", " ").strip())


def _plan_component_label(component: str) -> str:
    return FRIENDLY_PLAN_COMPONENT_LABELS.get(component, component.replace("_", " ").strip())


def _extend_unique(target: list[str], values: list[str]) -> None:
    for value in values:
        if value and value not in target:
            target.append(value)


def build_phase11_rule_execution(
    *,
    input_file: Path,
    approved_rule_package_file: Path,
    ftl_food_items_file: Path | None = None,
    kde_contracts_file: Path | None = None,
    exemption_rules_file: Path | None = None,
    plan_components_file: Path | None = None,
    inbound_files: tuple[Path, ...] = (),
) -> Phase11RuleExecutionPackage:
    phase10 = build_phase10_customer_evidence(input_file=input_file, ftl_food_items_file=ftl_food_items_file)
    phase10c = build_phase10c_cte_hardening(input_file=input_file, ftl_food_items_file=ftl_food_items_file)
    rule_package = json.loads(approved_rule_package_file.read_text(encoding="utf-8"))
    approved_obligations = _approved_obligations(rule_package)
    evidence_by_id = {record.evidence_id: record for record in phase10.evidence_records}
    event_by_id = {event.event_id: event for event in phase10.event_graph}
    hardened_by_event = {result.event_id: result for result in phase10c.production_event_results}
    # A farm map (21 CFR 1.1315(a)(4)) is only required of growers/harvesters. Detect a
    # farm/grower operation from the hardened CTEs so we don't flag a distributor/processor
    # for a "missing" farm map it never needs.
    is_farm_operation = any(
        cte in FARM_OPERATION_CTES
        for result in phase10c.production_event_results
        for cte in result.final_ctes
    )

    obligation_mappings = map_events_to_approved_obligations(
        hardened_results=phase10c.production_event_results,
        approved_obligations=approved_obligations,
        rule_package=rule_package,
    )
    kde_checks = check_kde_completeness(
        mappings=obligation_mappings,
        events=event_by_id,
        evidence_by_id=evidence_by_id,
        contracts_file=kde_contracts_file,
    )
    tlc_checks = check_tlc_lineage(
        mappings=obligation_mappings,
        events=event_by_id,
    )
    traceability_plan_checks = check_traceability_plan(
        phase10=phase10,
        approved_obligations=approved_obligations,
        rule_package=rule_package,
        is_farm_operation=is_farm_operation,
        plan_components_file=plan_components_file,
    )
    scope_exemption_checks = check_scope_and_exemptions(
        events=event_by_id,
        hardened_results=hardened_by_event,
        evidence_conflicts=phase10.evidence_conflicts,
        quality_report=phase10.quality_report.model_dump(mode="json") if phase10.quality_report else {},
    )
    scope_exemption_checks += check_exemption_claims(
        phase10=phase10, exemption_rules=_load_exemption_rules(exemption_rules_file)
    )
    records_readiness_checks = check_records_readiness(
        phase10=phase10,
        approved_obligations=approved_obligations,
    )
    sortable_export_checks = check_sortable_export_readiness(kde_checks)

    # --- Deep validation checks (lot integrity, GS1/overlays, FTL tiers, partner scorecard)
    from bellwether_backend.audit_engine.customer_evidence import _load_optional_json_list, _row_facts
    from bellwether_backend.audit_engine.gs1 import check_gs1_identifiers
    from bellwether_backend.audit_engine.lot_integrity import check_lot_integrity, compute_export_window
    from bellwether_backend.audit_engine.partner_scorecard import build_partner_scorecard, scorecard_summary_findings
    from bellwether_backend.audit_engine.scoping_report import build_scoping_report, build_scoping_stats
    from bellwether_backend.intelligence.ftl_tier_classifier import classify_products

    row_facts = _row_facts(phase10.evidence_records)
    export_window = compute_export_window(event_by_id)
    lot_integrity_checks = check_lot_integrity(events=event_by_id, row_facts=row_facts, export_window=export_window)
    gs1_checks = check_gs1_identifiers(entity_graph=phase10.entity_graph)
    ftl_items = _load_optional_json_list(ftl_food_items_file)
    ftl_products = [
        {
            "product_id": product.entity_id,
            "name": product.name,
            "declared_category": (product.attributes or {}).get("ftl_category"),
        }
        for product in phase10.entity_graph.products
    ]
    ftl_tier_results = classify_products(ftl_products, ftl_items) if ftl_products else {}
    partner_scorecard = build_partner_scorecard(
        events=event_by_id,
        entity_graph=phase10.entity_graph,
        row_facts=row_facts,
        lot_integrity_checks=lot_integrity_checks,
    )

    # Door-vs-database: diff supplier-provided inbound documents (ASN/EDI/BOL) against what
    # actually landed in the system of record.
    inbound_findings: list[dict[str, Any]] = []
    if inbound_files:
        from bellwether_backend.audit_engine.inbound_diff import diff_inbound_vs_erp

        for inbound_file in inbound_files:
            lines, label = _read_inbound_lines(inbound_file)
            if lines:
                inbound_findings.extend(
                    diff_inbound_vs_erp(
                        inbound_lines=lines,
                        events=event_by_id,
                        row_facts=row_facts,
                        source_label=label,
                    )
                )

    audit_findings = generate_audit_findings(
        kde_checks=kde_checks,
        tlc_checks=tlc_checks,
        traceability_plan_checks=traceability_plan_checks,
        scope_exemption_checks=scope_exemption_checks,
        records_readiness_checks=records_readiness_checks,
        sortable_export_checks=sortable_export_checks,
        approved_obligations=approved_obligations,
        lot_integrity_checks=lot_integrity_checks,
        gs1_checks=gs1_checks,
        ftl_tier_results=ftl_tier_results,
        partner_summary_findings=scorecard_summary_findings(partner_scorecard) + inbound_findings,
    )
    exception_queue = generate_exception_queue(audit_findings)
    export_package = build_fda_style_export_package(
        rule_package=rule_package,
        events=event_by_id,
        hardened_results=phase10c.production_event_results,
        kde_checks=kde_checks,
        audit_findings=audit_findings,
        sortable_export_checks=sortable_export_checks,
    )
    summary = _summary(
        rule_package=rule_package,
        obligation_mappings=obligation_mappings,
        kde_checks=kde_checks,
        tlc_checks=tlc_checks,
        traceability_plan_checks=traceability_plan_checks,
        scope_exemption_checks=scope_exemption_checks,
        records_readiness_checks=records_readiness_checks,
        sortable_export_checks=sortable_export_checks,
        audit_findings=audit_findings,
        exception_queue=exception_queue,
        export_package=export_package,
    )
    mapping_plan_summary = None
    if phase10.mapping_plan:
        mapping_plan_summary = {
            "generatedBy": phase10.mapping_plan.get("generated_by"),
            "sheetKinds": {
                name: sheet.get("record_kind")
                for name, sheet in (phase10.mapping_plan.get("sheet_plans") or {}).items()
            },
        }
    scoping_stats = build_scoping_stats(
        events=event_by_id,
        ftl_tier_results=ftl_tier_results,
        partner_scorecard=partner_scorecard,
        kde_checks=kde_checks,
        lot_integrity_checks=lot_integrity_checks,
        audit_findings=audit_findings,
        export_window=export_window,
        mapping_plan_summary=mapping_plan_summary,
    )
    scoping_report = build_scoping_report(stats=scoping_stats)
    summary["scoping"] = scoping_stats
    summary["lotIntegrityStatusCounts"] = dict(sorted(Counter(f"{c.check_type}:{c.status}" for c in lot_integrity_checks).items()))
    summary["gs1InvalidCount"] = sum(1 for c in gs1_checks if not c.valid_check_digit)
    return Phase11RuleExecutionPackage(
        generated_at=GENERATED_AT,
        summary=summary,
        obligation_mappings=obligation_mappings,
        kde_checks=kde_checks,
        tlc_checks=tlc_checks,
        traceability_plan_checks=traceability_plan_checks,
        scope_exemption_checks=scope_exemption_checks,
        records_readiness_checks=records_readiness_checks,
        sortable_export_checks=sortable_export_checks,
        audit_findings=audit_findings,
        exception_queue=exception_queue,
        export_package=export_package,
        lot_integrity_checks=lot_integrity_checks,
        gs1_checks=gs1_checks,
        ftl_tier_results=ftl_tier_results,
        partner_scorecard=partner_scorecard,
        scoping_report=scoping_report,
        mapping_plan=phase10.mapping_plan,
    )


def write_phase11_rule_execution_artifacts(package: Phase11RuleExecutionPackage, output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = {
        "summary": output_dir / "phase11-summary.json",
        "obligationMapping": output_dir / "phase11-obligation-mapping.json",
        "kdeCompleteness": output_dir / "phase11-kde-completeness-results.json",
        "tlcLineage": output_dir / "phase11-tlc-lineage-results.json",
        "traceabilityPlan": output_dir / "phase11-traceability-plan-results.json",
        "scopeExemption": output_dir / "phase11-scope-exemption-results.json",
        "recordsReadiness": output_dir / "phase11-records-readiness-results.json",
        "sortableExportReadiness": output_dir / "phase11-sortable-export-readiness.json",
        "auditFindings": output_dir / "phase11-audit-findings.json",
        "exceptionQueue": output_dir / "phase11-exception-queue.json",
        "exportPackage": output_dir / "phase11-export-package.json",
        "exportWorkbook": output_dir / "phase11-fda-style-export-package.xlsx",
        "lotIntegrity": output_dir / "phase11-lot-integrity-results.json",
        "gs1Results": output_dir / "phase11-gs1-results.json",
        "ftlTierResults": output_dir / "phase11-ftl-tier-results.json",
        "partnerScorecard": output_dir / "phase11-partner-scorecard.json",
        "scopingReport": output_dir / "phase11-scoping-report.json",
        "workbookMappingPlan": output_dir / "phase11-workbook-mapping-plan.json",
    }
    _write_json(outputs["summary"], package.summary)
    _write_json(outputs["obligationMapping"], [item.model_dump(mode="json") for item in package.obligation_mappings])
    _write_json(outputs["kdeCompleteness"], [item.model_dump(mode="json") for item in package.kde_checks])
    _write_json(outputs["tlcLineage"], [item.model_dump(mode="json") for item in package.tlc_checks])
    _write_json(outputs["traceabilityPlan"], [item.model_dump(mode="json") for item in package.traceability_plan_checks])
    _write_json(outputs["scopeExemption"], [item.model_dump(mode="json") for item in package.scope_exemption_checks])
    _write_json(outputs["recordsReadiness"], [item.model_dump(mode="json") for item in package.records_readiness_checks])
    _write_json(outputs["sortableExportReadiness"], [item.model_dump(mode="json") for item in package.sortable_export_checks])
    _write_json(outputs["auditFindings"], [item.model_dump(mode="json") for item in package.audit_findings])
    _write_json(outputs["exceptionQueue"], [item.model_dump(mode="json") for item in package.exception_queue])
    export_payload = package.export_package.model_copy(update={"workbook_file": str(outputs["exportWorkbook"])}).model_dump(mode="json")
    _write_json(outputs["exportPackage"], export_payload)
    _write_export_workbook(outputs["exportWorkbook"], package.export_package, package.audit_findings)
    _write_json(outputs["lotIntegrity"], [item.model_dump(mode="json") for item in package.lot_integrity_checks])
    _write_json(outputs["gs1Results"], [item.model_dump(mode="json") for item in package.gs1_checks])
    _write_json(outputs["ftlTierResults"], package.ftl_tier_results)
    _write_json(outputs["partnerScorecard"], package.partner_scorecard)
    _write_json(outputs["scopingReport"], package.scoping_report)
    _write_json(outputs["workbookMappingPlan"], package.mapping_plan or {})
    return {key: str(path) for key, path in outputs.items()}


def map_events_to_approved_obligations(
    *,
    hardened_results: list[MultiSignalCteResult],
    approved_obligations: dict[str, dict[str, Any]],
    rule_package: dict[str, Any],
) -> list[EventObligationMapping]:
    mappings: list[EventObligationMapping] = []
    for result in hardened_results:
        for cte in result.final_ctes:
            for obligation in approved_obligations.values():
                applies = obligation.get("applies_to_ctes") or []
                if cte not in applies and "other" not in applies:
                    continue
                if "other" in applies:
                    continue
                if not _is_specific_cte_obligation(obligation, cte):
                    continue
                mappings.append(_mapping(result.event_id, cte, obligation, rule_package, len(mappings) + 1))
    other_ctes = {"records_readiness", "sortable_export"}
    for result in hardened_results:
        if result.final_ctes:
            for obligation in approved_obligations.values():
                if "other" in (obligation.get("applies_to_ctes") or []):
                    for cte in other_ctes:
                        mappings.append(_mapping(result.event_id, cte, obligation, rule_package, len(mappings) + 1))
    return mappings


def check_kde_completeness(
    *,
    mappings: list[EventObligationMapping],
    events: dict[str, CustomerEventNode],
    evidence_by_id: dict[str, CustomerEvidenceRecord],
    contracts_file: Path | None = None,
) -> list[KdeCompletenessCheck]:
    contracts = _load_kde_check_contracts(contracts_file)
    # Promote the customer's 07_KDE_Values rows (each declares its own field key + value) into
    # per-event facts so KDEs the workbook DOES carry — source references, harvest origin —
    # actually feed the completeness check instead of being perpetually "not captured".
    promoted_kde_facts = _promote_kde_value_rows(evidence_by_id)
    checks: list[KdeCompletenessCheck] = []
    for mapping in mappings:
        cte_contract = contracts.get(mapping.cte)
        if not cte_contract:
            continue
        section = cte_contract.get("citation_section", "")
        event = events[mapping.event_id]
        facts = _event_facts(event, evidence_by_id)
        _merge_promoted_kde_facts(facts, promoted_kde_facts, event.event_id)
        for result in evaluate_kde_contract_facts(cte_contract=cte_contract, facts=facts):
            checks.append(
                KdeCompletenessCheck(
                    check_id=f"phase11-kde-{len(checks) + 1:04d}",
                    event_id=mapping.event_id,
                    cte=mapping.cte,
                    field_key=result["kde"],
                    status=result["status"],
                    expected_reason=result["expected_reason"],
                    evidence_ids=result["evidence_ids"],
                    observed_values=result["observed_values"],
                    approved_obligation_id=mapping.approved_obligation_id,
                )
            )
    return checks


def evaluate_kde_contract_facts(*, cte_contract: dict[str, Any], facts: dict[str, list[str]]) -> list[dict[str, Any]]:
    """Evaluate one CTE's KDE contract against a facts dict — the reusable core shared by the
    workbook audit and the pre-receipt inbound validation endpoint.

    ``facts``: {canonical_slug: [values], "evidence:<slug>": [evidence ids]}.
    """
    section = cte_contract.get("citation_section", "")
    results: list[dict[str, Any]] = []
    for kde in cte_contract["kdes"]:
        kde_key = str(kde["kde"])
        satisfied_by = kde.get("satisfied_by") or []
        label = kde.get("label", kde_key)
        # A placeholder ("UNKNOWN", "N/A", "TBD", ...) is NOT a real value — it must not
        # count as a present KDE, or a non-compliant record would pass (a false pass).
        # satisfied_by is "any of": the KDE is met if ANY of those fields carries a value.
        real_by_field = {field: sorted({v for v in facts.get(field, []) if _is_answered(v)}) for field in satisfied_by}
        values = sorted({v for vals in real_by_field.values() for v in vals})
        evidence_ids = sorted({ev for field in satisfied_by for ev in facts.get(f"evidence:{field}", [])})
        if not satisfied_by:
            # Required by FSMA but the parser does not extract this field yet. Tracked
            # honestly (not failed, not silently dropped) until parser support is added.
            status = "not_captured"
        elif len(satisfied_by) == 1 and len(real_by_field[satisfied_by[0]]) > 1:
            status = "conflicting"
        elif values:
            status = "present"
        elif kde.get("requirement") == "conditional":
            status = "not_applicable"
        else:
            status = "missing"
        results.append(
            {
                "kde": kde_key,
                "label": label,
                "status": status,
                "requirement": kde.get("requirement", "required"),
                "severity": kde.get("severity", "medium"),
                "expected_reason": f"{label} ({section})" if section else label,
                "citation_section": section,
                "observed_values": values,
                "evidence_ids": evidence_ids,
            }
        )
    return results


def check_tlc_lineage(
    *,
    mappings: list[EventObligationMapping],
    events: dict[str, CustomerEventNode],
) -> list[TlcLineageCheck]:
    # Lots that exist as a real receipt/output anywhere in these records. A transformation
    # input must trace back to one of these, otherwise the chain is broken even though a
    # lot code is "present" — presence alone is not linkage.
    upstream_lots = {
        value.strip().lower()
        for event in events.values()
        for value in (_real_value(event.lot_or_tlc), _real_value(event.output_lot_or_tlc))
        if value
    }

    checks: list[TlcLineageCheck] = []
    for mapping in mappings:
        if mapping.cte not in {"initial_packing", "first_land_based_receiving", "transformation", "shipping", "receiving"}:
            continue
        event = events[mapping.event_id]
        lot = _real_value(event.lot_or_tlc)
        source = _real_value(event.source_lot_or_tlc)
        output = _real_value(event.output_lot_or_tlc)
        if mapping.cte == "transformation":
            if not (output and source):
                status = "gap"
                reason = "Transformation must record both an incoming (source) and a new (output) traceability lot code."
            elif source.strip().lower() not in upstream_lots:
                status = "gap"
                reason = "Transformation input lot code does not trace to any upstream receiving/packing record — the lineage is broken."
            else:
                status = "linked"
                reason = "Input and output lot codes are present and the input lot traces to an upstream event."
        elif mapping.cte in {"initial_packing", "first_land_based_receiving"}:
            status = "linked" if (lot or output) else "gap"
            reason = "A new traceability lot code must be assigned at this event." if status == "gap" else "Traceability lot code assigned."
        else:
            status = "linked" if lot else "gap"
            reason = "A traceability lot code must be present on this shipping/receiving record." if status == "gap" else "Traceability lot code present."
        checks.append(
            TlcLineageCheck(
                check_id=f"phase11-tlc-{len(checks) + 1:04d}",
                event_id=mapping.event_id,
                cte=mapping.cte,
                status=status,
                lot_or_tlc=event.lot_or_tlc,
                source_lot_or_tlc=event.source_lot_or_tlc,
                output_lot_or_tlc=event.output_lot_or_tlc,
                evidence_ids=event.evidence_ids,
                approved_obligation_id=mapping.approved_obligation_id,
                reason=reason,
            )
        )
    return checks


def check_traceability_plan(
    *,
    phase10: Phase10CustomerEvidencePackage,
    approved_obligations: dict[str, dict[str, Any]],
    rule_package: dict[str, Any],
    is_farm_operation: bool = False,
    plan_components_file: Path | None = None,
) -> list[TraceabilityPlanCheck]:
    obligation = _find_obligation_by_cte(approved_obligations, "traceability_plan")
    plan_records = [record for record in phase10.evidence_records if record.sheet_name == "04_Traceability_Plan"]

    # Reconstruct plan rows (item -> answer). A component is only "present" if its row
    # exists AND its answer cell is actually filled in. The previous logic matched on the
    # item label alone, so a blank answer (e.g. a missing TLC-assignment procedure) was
    # silently counted as present. For a compliance tool, a missed gap is worse than a
    # duplicate, so blank/placeholder answers must surface as gaps.
    rows: "OrderedDict[Any, dict[str, Any]]" = OrderedDict()
    for record in plan_records:
        row = rows.setdefault(record.row_number, {"item": None, "answer": "", "evidence_ids": []})
        if record.field_key == "traceability_plan_item":
            row["item"] = record.normalized_value.strip().lower()
        elif record.field_key == "traceability_plan_answer":
            row["answer"] = record.normalized_value.strip()
        _extend_unique(row["evidence_ids"], [record.evidence_id])

    components = _load_plan_components(plan_components_file)
    checks: list[TraceabilityPlanCheck] = []
    for spec in components:
        component = spec["component"]
        terms = spec.get("match_terms", [])
        farm_only = bool(spec.get("farm_only"))
        matching_rows = [row for row in rows.values() if row["item"] and any(term in row["item"] for term in terms)]
        answered_rows = [row for row in matching_rows if _is_answered(row["answer"])]
        evidence_ids: list[str] = []
        for row in matching_rows:
            _extend_unique(evidence_ids, row["evidence_ids"])
        if answered_rows:
            status = "present"
            reason = f"Traceability plan component expected by approved obligation in {rule_package['package_id']}."
        elif farm_only and not is_farm_operation:
            status = "not_applicable"
            reason = "Farm map is required only for farms/growers (21 CFR 1.1315(a)(4)); not applicable to this operation."
        elif matching_rows:
            status = "missing"
            reason = "Traceability plan lists this component but the answer is blank or undetermined."
        else:
            status = "missing"
            reason = "Traceability plan does not document this required component."
        checks.append(
            TraceabilityPlanCheck(
                check_id=f"phase11-plan-{len(checks) + 1:04d}",
                component=component,
                status=status,
                evidence_ids=evidence_ids,
                approved_obligation_id=obligation["obligation_id"],
                reason=reason,
            )
        )
    return checks


_PLACEHOLDER_VALUES = {
    "",
    "unknown",
    "n/a",
    "n.a.",
    "na",
    "tbd",
    "tbc",
    "none",
    "null",
    "-",
    "--",
    "pending",
    "not provided",
    "not available",
    "unassigned",
    "missing",
}


def _is_answered(value: str) -> bool:
    return value.strip().lower() not in _PLACEHOLDER_VALUES


def _real_value(value: str | None) -> str | None:
    """A customer-supplied value, or None if it's blank/a placeholder ("UNKNOWN", "N/A", ...).
    A placeholder must never count as a present KDE or a linked TLC — that would be a false pass."""
    if value is None:
        return None
    return value if _is_answered(value) else None


def _scope_review_reason(event: CustomerEventNode, hardened: MultiSignalCteResult | None) -> str:
    product = (getattr(event, "product_name", "") or "").strip() or "This record"
    questions = list(getattr(hardened, "reviewer_questions", []) or []) if hardened else []
    if questions:
        return f"{product} needs a scope/exemption review: {questions[0]}"
    return f"{product} needs a food-scope review before this record can produce a final finding."


def check_scope_and_exemptions(
    *,
    events: dict[str, CustomerEventNode],
    hardened_results: dict[str, MultiSignalCteResult],
    evidence_conflicts: list[Any],
    quality_report: dict[str, Any],
) -> list[ScopeExemptionCheck]:
    checks: list[ScopeExemptionCheck] = []
    for event_id, event in events.items():
        hardened = hardened_results.get(event_id)
        if event.food_form.review_required or (hardened and hardened.reviewer_questions):
            checks.append(
                ScopeExemptionCheck(
                    check_id=f"phase11-scope-{len(checks) + 1:04d}",
                    event_id=event_id,
                    check_type="scope_or_exemption_uncertainty",
                    status="needs_review",
                    reason=_scope_review_reason(event, hardened),
                    evidence_ids=event.evidence_ids,
                )
            )
    if evidence_conflicts:
        checks.append(
            ScopeExemptionCheck(
                check_id=f"phase11-scope-{len(checks) + 1:04d}",
                event_id=None,
                check_type="evidence_conflicts",
                status="needs_review",
                reason="Customer evidence conflicts must be resolved before final execution.",
                evidence_ids=[],
            )
        )
    if quality_report.get("quality_gate") not in {None, "pass"}:
        checks.append(
            ScopeExemptionCheck(
                check_id=f"phase11-scope-{len(checks) + 1:04d}",
                event_id=None,
                check_type="customer_evidence_quality_gate",
                status="needs_review",
                reason=f"Customer evidence quality gate is {quality_report.get('quality_gate')}.",
                evidence_ids=[],
            )
        )
    return checks


def _normalize_match_token(value: str) -> str:
    """Lower-case, collapse non-alphanumerics to single spaces — so 'small_producer',
    'Small Producer', and 'small-producer' all compare equal."""
    return " ".join("".join(ch if ch.isalnum() else " " for ch in str(value).lower()).split())


def _match_exemption_rule(claim_type: str, exemption_rules: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Match a customer's claimed exemption type to an approved 21 CFR 1.1305 rule. Tolerant of
    both card schemas: the bundled rules (canonical snake_case exemption_type + aliases) and the
    approved Supabase cards (prose exemption_type, exemption_rule_id, no aliases). Matches on a
    normalized exemption_type/alias/id, then on token containment as a last resort."""
    needle = _normalize_match_token(claim_type)
    if not needle:
        return None
    for rule in exemption_rules:
        candidates = [rule.get("exemption_type", ""), rule.get("exemption_rule_id", "")]
        candidates.extend(rule.get("aliases", []) or [])
        normalized = [_normalize_match_token(c) for c in candidates if c]
        if any(needle == c for c in normalized):
            return rule
    # Last resort: the claim text contains (or is contained by) a known type/alias token string.
    for rule in exemption_rules:
        candidates = [rule.get("exemption_type", "")] + list(rule.get("aliases", []) or [])
        for candidate in candidates:
            token = _normalize_match_token(candidate)
            if token and (token in needle or needle in token):
                return rule
    return None


def _exemption_evidence_text(rule: dict[str, Any]) -> str:
    """The documentation a customer must show, across both card schemas."""
    value = rule.get("evidence_required") or rule.get("documentation_needed") or rule.get("eligibility_condition")
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(item) for item in value if item)
    return str(value).strip() if value else "supporting eligibility records"


def _exemption_label(rule: dict[str, Any], fallback: str) -> str:
    raw = str(rule.get("exemption_type") or fallback).replace("_", " ").strip()
    return raw[:1].upper() + raw[1:] if raw else fallback


def check_exemption_claims(
    *,
    phase10: Phase10CustomerEvidencePackage,
    exemption_rules: list[dict[str, Any]],
) -> list[ScopeExemptionCheck]:
    """Evaluate the customer's claimed exemptions (sheet 10_Exemptions_Claims) against the
    approved 21 CFR 1.1305 exemption rules.

    Conservative by design: a claim is NEVER allowed to auto-suppress a CTE/KDE obligation.
    Every claim is surfaced for human confirmation. A claim with supporting evidence becomes a
    'review/confirm' item; a claim with no evidence is 'not_determined' (we neither grant nor
    deny it); an unrecognized claim type is flagged for reviewer judgment. Telling a customer
    they are exempt when they are not is the worst possible error, so the engine only ever
    raises the question — a human grants the exemption."""
    claim_records = [r for r in phase10.evidence_records if r.sheet_name == "10_Exemptions_Claims"]
    rows: "OrderedDict[Any, dict[str, Any]]" = OrderedDict()
    for record in claim_records:
        row = rows.setdefault(
            record.row_number,
            {"claim_id": None, "claim_type": None, "claimed_by": None, "evidence": "", "evidence_ids": []},
        )
        if record.field_key == "exemption_claim_id":
            row["claim_id"] = record.normalized_value.strip()
        elif record.field_key == "exemption_claim_type":
            row["claim_type"] = record.normalized_value.strip()
        elif record.field_key == "exemption_claimed_by":
            row["claimed_by"] = record.normalized_value.strip()
        elif record.field_key == "exemption_evidence_provided":
            row["evidence"] = record.normalized_value.strip()
        _extend_unique(row["evidence_ids"], [record.evidence_id])

    checks: list[ScopeExemptionCheck] = []
    for row in rows.values():
        claim_type = (row["claim_type"] or "").strip()
        if not claim_type:
            continue
        claimant = row["claimed_by"] or "the operation"
        rule = _match_exemption_rule(claim_type, exemption_rules)
        has_evidence = row["evidence"].strip().lower() in {"yes", "y", "true", "1", "provided", "attached"}
        if rule is None:
            status = "needs_review"
            reason = (
                f"{claimant} claims a '{claim_type}' exemption that could not be matched automatically to an "
                "approved 21 CFR 1.1305 exemption rule. A reviewer must confirm whether any exemption applies "
                "before the related records are treated as not required."
            )
        else:
            label = _exemption_label(rule, claim_type)
            effect = str(rule.get("effect") or "review required")
            if has_evidence:
                status = "needs_review"
                reason = (
                    f"{claimant} claims the {label} exemption ({effect}) and provided supporting evidence. "
                    "Confirm the evidence before relying on it — Bellwether does not auto-grant exemptions."
                )
            else:
                status = "not_determined"
                reason = (
                    f"{claimant} claims the {label} exemption ({effect}) but provided no supporting evidence "
                    f"({_exemption_evidence_text(rule)}). The exemption is not determined; the underlying records "
                    "remain required until eligibility is documented and confirmed."
                )
        checks.append(
            ScopeExemptionCheck(
                check_id=f"phase11-exemption-{len(checks) + 1:04d}",
                event_id=None,
                check_type="exemption_claim",
                status=status,
                reason=reason,
                evidence_ids=row["evidence_ids"],
            )
        )
    return checks


def check_records_readiness(
    *,
    phase10: Phase10CustomerEvidencePackage,
    approved_obligations: dict[str, dict[str, Any]],
) -> list[RecordsReadinessCheck]:
    records_obligation = _find_obligation_containing(approved_obligations, "RECORDS-MAINTENANCE")
    request_obligation = _find_obligation_containing(approved_obligations, "FDA-REQUEST")
    export_obligation = _find_obligation_containing(approved_obligations, "SORTABLE-SPREADSHEET")
    evidence_records = phase10.evidence_records
    checks = [
        RecordsReadinessCheck(
            check_id="phase11-records-0001",
            check_type="legible_linked_records",
            status="present" if evidence_records and not phase10.evidence_conflicts else "needs_review",
            reason="Parsed evidence has cell-level lineage and no unresolved evidence conflicts." if not phase10.evidence_conflicts else "Evidence conflicts require review.",
            evidence_ids=[record.evidence_id for record in evidence_records[:50]],
            approved_obligation_id=records_obligation["obligation_id"],
        ),
        RecordsReadinessCheck(
            check_id="phase11-records-0002",
            check_type="fda_24_hour_response_readiness",
            status="present" if phase10.quality_report and phase10.quality_report.quality_gate == "pass" else "needs_review",
            reason="Evidence package is parseable and linked for FDA request workflow." if phase10.quality_report and phase10.quality_report.quality_gate == "pass" else "Evidence quality gate blocks FDA response readiness.",
            evidence_ids=[record.evidence_id for record in evidence_records[:50]],
            approved_obligation_id=request_obligation["obligation_id"],
        ),
        RecordsReadinessCheck(
            check_id="phase11-records-0003",
            check_type="sortable_export_source_data",
            status="present" if phase10.event_graph else "missing",
            reason="Normalized event graph exists for sortable export population.",
            evidence_ids=[evidence_id for event in phase10.event_graph for evidence_id in event.evidence_ids[:5]],
            approved_obligation_id=export_obligation["obligation_id"],
        ),
    ]
    return checks


def check_sortable_export_readiness(kde_checks: list[KdeCompletenessCheck]) -> list[SortableExportReadinessCheck]:
    by_event_cte: dict[tuple[str, str], list[KdeCompletenessCheck]] = defaultdict(list)
    for check in kde_checks:
        by_event_cte[(check.event_id, check.cte)].append(check)
    results: list[SortableExportReadinessCheck] = []
    for (event_id, cte), checks in sorted(by_event_cte.items()):
        missing = [check.field_key for check in checks if check.status in {"missing", "conflicting"}]
        populated = [check.field_key for check in checks if check.status == "present"]
        results.append(
            SortableExportReadinessCheck(
                check_id=f"phase11-export-{len(results) + 1:04d}",
                event_id=event_id,
                cte=cte,
                status="ready" if not missing else "blocked",
                missing_fields=missing,
                populated_fields=populated,
                approved_obligation_id=checks[0].approved_obligation_id,
            )
        )
    return results


def generate_audit_findings(
    *,
    kde_checks: list[KdeCompletenessCheck],
    tlc_checks: list[TlcLineageCheck],
    traceability_plan_checks: list[TraceabilityPlanCheck],
    scope_exemption_checks: list[ScopeExemptionCheck],
    records_readiness_checks: list[RecordsReadinessCheck],
    sortable_export_checks: list[SortableExportReadinessCheck],
    approved_obligations: dict[str, dict[str, Any]],
    lot_integrity_checks: list[Any] = (),
    gs1_checks: list[Any] = (),
    ftl_tier_results: dict[str, dict[str, Any]] | None = None,
    partner_summary_findings: list[dict[str, Any]] = (),
) -> list[AuditFinding]:
    findings: list[AuditFinding] = []
    fallback_obligation = next(iter(approved_obligations.values()))

    # --- Event-scoped gaps: one finding per record, root cause grouped ---
    # A missing TLC commonly surfaces as both a KDE-completeness gap and a TLC-lineage
    # gap on the same event. We merge all KDE/TLC gaps for one event into a single
    # record-level finding so the partner sees one real problem, not several copies.
    event_groups: "OrderedDict[str, dict[str, Any]]" = OrderedDict()

    def _group_for(event_id: str, cte: str, obligation_id: str | None) -> dict[str, Any]:
        group = event_groups.get(event_id)
        if group is None:
            group = {
                "event_id": event_id,
                "cte": cte,
                "missing_fields": [],
                "tlc_reasons": [],
                "evidence_ids": [],
                "obligation_id": obligation_id,
                "tlc_root": False,
            }
            event_groups[event_id] = group
        if not group.get("obligation_id") and obligation_id:
            group["obligation_id"] = obligation_id
        return group

    for check in kde_checks:
        # Only real gaps become findings. "present" passes; "not_applicable" (conditional KDE
        # that doesn't apply) and "not_captured" (FSMA-required but not yet parser-extracted)
        # are not customer-facing failures.
        if check.status in {"present", "not_applicable", "not_captured"}:
            continue
        group = _group_for(check.event_id, check.cte, check.approved_obligation_id)
        if check.field_key not in group["missing_fields"]:
            group["missing_fields"].append(check.field_key)
        _extend_unique(group["evidence_ids"], check.evidence_ids)
        if check.field_key in TLC_FIELD_KEYS:
            group["tlc_root"] = True

    for check in tlc_checks:
        if check.status == "linked":
            continue
        group = _group_for(check.event_id, check.cte, check.approved_obligation_id)
        if check.reason not in group["tlc_reasons"]:
            group["tlc_reasons"].append(check.reason)
        _extend_unique(group["evidence_ids"], check.evidence_ids)
        group["tlc_root"] = True

    # The same root cause repeated across many records is ONE systemic problem, not N
    # separate findings (e.g. every transformation row in a two-sheet export missing its
    # source linkage, or every landing record missing harvest info). Cluster identical gap
    # signatures; clusters above the threshold collapse into a single systemic finding.
    SYSTEMIC_THRESHOLD = 4
    clusters: "OrderedDict[tuple[Any, ...], list[dict[str, Any]]]" = OrderedDict()
    for group in event_groups.values():
        signature = (
            group["cte"],
            tuple(sorted(group["missing_fields"])),
            tuple(sorted(group["tlc_reasons"])),
            bool(group["tlc_root"]),
        )
        clusters.setdefault(signature, []).append(group)

    for signature, grouped in clusters.items():
        cte, missing_fields_key, tlc_reasons_key, is_tlc_root = signature
        missing_fields = list(missing_fields_key)
        tlc_reasons = list(tlc_reasons_key)
        high = is_tlc_root or any(field in HIGH_SEVERITY_FIELD_KEYS for field in missing_fields)
        finding_type = "tlc_lineage" if is_tlc_root else "kde_completeness"
        if len(grouped) >= SYSTEMIC_THRESHOLD:
            evidence_ids: list[str] = []
            for group in grouped[:10]:
                _extend_unique(evidence_ids, group["evidence_ids"][:5])
            obligation = approved_obligations.get(grouped[0]["obligation_id"]) or fallback_obligation
            base_message = _record_gap_message(cte, missing_fields, is_tlc_root).rstrip(".")
            findings.append(
                _finding(
                    findings,
                    event_id=None,
                    cte=cte,
                    severity="high" if high else "medium",
                    status="gap",
                    finding_type=finding_type,
                    message=(
                        f"Systemic gap across {len(grouped)} {_cte_label(cte).lower()} records: "
                        f"{base_message[0].lower()}{base_message[1:]}. This repeats on every affected "
                        "record, which points at the source system/template rather than data entry."
                    ),
                    obligation=obligation,
                    evidence_ids=evidence_ids,
                    confidence=0.9 if is_tlc_root else 0.88,
                    sub_issues=_record_sub_issues(missing_fields, tlc_reasons)
                    + [f"Affected records: {len(grouped)} (e.g. {', '.join(str(g['event_id']) for g in grouped[:5])})"],
                    affected_fields=missing_fields,
                )
            )
            continue
        for group in grouped:
            obligation = approved_obligations.get(group["obligation_id"]) or fallback_obligation
            findings.append(
                _finding(
                    findings,
                    event_id=group["event_id"],
                    cte=group["cte"],
                    severity="high" if high else "medium",
                    status="gap",
                    finding_type=finding_type,
                    message=_record_gap_message(group["cte"], group["missing_fields"], is_tlc_root),
                    obligation=obligation,
                    evidence_ids=group["evidence_ids"],
                    confidence=0.9 if is_tlc_root else 0.88,
                    sub_issues=_record_sub_issues(group["missing_fields"], group["tlc_reasons"]),
                    affected_fields=list(group["missing_fields"]),
                )
            )

    # --- Traceability plan: one finding listing all missing components ---
    missing_plan = [check for check in traceability_plan_checks if check.status not in {"present", "not_applicable"}]
    if missing_plan:
        plan_evidence: list[str] = []
        for check in missing_plan:
            _extend_unique(plan_evidence, check.evidence_ids)
        components = [check.component for check in missing_plan]
        findings.append(
            _finding(
                findings,
                event_id=None,
                cte="traceability_plan",
                severity="medium",
                status="gap",
                finding_type="traceability_plan",
                message=(
                    f"Traceability plan is incomplete: {len(components)} required "
                    f"component{'s' if len(components) != 1 else ''} not documented."
                ),
                obligation=approved_obligations.get(missing_plan[0].approved_obligation_id) or fallback_obligation,
                evidence_ids=plan_evidence,
                confidence=0.86,
                sub_issues=[f"Missing {_plan_component_label(component)}" for component in components],
                affected_fields=list(components),
            )
        )

    # --- Scope / exemption review items: grouped per record + check type ---
    scope_groups: "OrderedDict[tuple[str, str], dict[str, Any]]" = OrderedDict()
    for check in scope_exemption_checks:
        if check.check_type == "customer_evidence_quality_gate":
            continue
        key = (check.event_id or "__global__", check.check_type)
        group = scope_groups.get(key)
        if group is None:
            group = {"event_id": check.event_id, "check_type": check.check_type, "reasons": [], "evidence_ids": []}
            scope_groups[key] = group
        if check.reason not in group["reasons"]:
            group["reasons"].append(check.reason)
        _extend_unique(group["evidence_ids"], check.evidence_ids)

    for group in scope_groups.values():
        reasons = group["reasons"]
        message = reasons[0] if len(reasons) == 1 else f"{len(reasons)} items need reviewer judgment for this record."
        findings.append(
            _finding(
                findings,
                event_id=group["event_id"],
                cte=None,
                severity="medium",
                status="needs_review",
                finding_type=group["check_type"],
                message=message,
                obligation=fallback_obligation,
                evidence_ids=group["evidence_ids"],
                confidence=0.68,
                sub_issues=list(reasons),
            )
        )

    # --- Lot & lineage integrity findings (deterministic, cited; systemic patterns rolled up) ---
    lot_clusters: "OrderedDict[tuple[str, str], list[Any]]" = OrderedDict()
    for check in lot_integrity_checks:
        if check.status in {"linked", "pass"}:
            continue
        lot_clusters.setdefault((check.check_type, check.status), []).append(check)
    for (check_type, status), checks in lot_clusters.items():
        example = checks[0]
        is_best_practice = getattr(example, "basis", "regulation") == "best_practice"
        citation = (
            _best_practice_citation()
            if is_best_practice
            else _regulation_citation(example.citation_section)
        )
        requirement_source = "best_practice" if is_best_practice else "fda_rule"
        suffix = (
            " [Recall-readiness data-quality check - not itself a Subpart S recordkeeping requirement.]"
            if is_best_practice
            else ""
        )
        if len(checks) >= 4:
            lots = [check.lot for check in checks if check.lot]
            evidence_ids: list[str] = []
            for check in checks[:10]:
                _extend_unique(evidence_ids, list(check.evidence_ids)[:3])
            findings.append(
                _finding(
                    findings,
                    event_id=None,
                    cte=example.cte,
                    severity=max((check.severity for check in checks), key=lambda s: s == "high"),
                    status=status,
                    finding_type=f"lot_{check_type}",
                    message=(
                        f"{len(checks)} lots share the same lot-integrity issue ({check_type.replace('_', ' ')}). "
                        f"Example: {example.reason}{suffix}"
                    ),
                    obligation=fallback_obligation,
                    evidence_ids=evidence_ids,
                    confidence=0.9,
                    source_citation_override=citation,
                    requirement_source=requirement_source,
                    sub_issues=[f"Lot {lot}" for lot in lots[:15]] + ([f"...and {len(lots) - 15} more"] if len(lots) > 15 else []),
                    affected_fields=["traceability_lot_code"],
                )
            )
            continue
        for check in checks:
            findings.append(
                _finding(
                    findings,
                    event_id=check.event_id,
                    cte=check.cte,
                    severity=check.severity,
                    status=check.status,
                    finding_type=f"lot_{check.check_type}",
                    message=f"{check.reason}{suffix}",
                    obligation=fallback_obligation,
                    evidence_ids=list(check.evidence_ids),
                    confidence=0.9,
                    source_citation_override=citation,
                    requirement_source=requirement_source,
                    affected_fields=["traceability_lot_code"],
                )
            )

    # --- FTL declared-vs-inferred mismatches (the "escaping traceability" headline) ---
    for product_id, result in sorted((ftl_tier_results or {}).items()):
        if not result.get("mismatch"):
            continue
        tier_label = "on the Food Traceability List" if result.get("tier") == "definite_on" else "potentially on the Food Traceability List"
        findings.append(
            _finding(
                findings,
                event_id=None,
                cte=None,
                severity="high" if result.get("tier") == "definite_on" else "medium",
                status="needs_review",
                finding_type="ftl_declared_mismatch",
                message=(
                    f"Product {product_id} is declared \"{result.get('declared_category') or 'not on FTL'}\" "
                    f"but its description reads as {tier_label} ({result.get('reasoning', '').strip()}) - "
                    "if so, its events are silently escaping FSMA 204 KDE requirements."
                ),
                obligation=fallback_obligation,
                evidence_ids=[],
                confidence=0.75,
                source_citation_override=_regulation_citation("21 CFR 1.1305"),
                affected_fields=["ftl_category"],
            )
        )

    # --- GS1 / retailer-overlay findings ---
    for check in gs1_checks:
        if check.valid_check_digit:
            continue
        overlay = check.requirement_source == "customer_requirement"
        findings.append(
            _finding(
                findings,
                event_id=None,
                cte=None,
                severity="medium",
                status="needs_review",
                finding_type="gs1_requirement" if overlay else "gs1_identifier",
                message=check.reason,
                obligation=fallback_obligation,
                evidence_ids=list(check.evidence_ids),
                confidence=0.85,
                source_citation_override=(
                    {"sourceType": "customer_requirement", "citation_anchor": f"{check.retailer} supplier requirements", "section_ref": f"{check.retailer} GS1 mandate"}
                    if overlay
                    else _regulation_citation("21 CFR 1.1340")
                ),
                requirement_source=check.requirement_source,
                affected_fields=["product_id" if check.entity_type == "product" else "location_id"],
            )
        )

    # --- Partner scorecard summary findings ---
    for summary_finding in partner_summary_findings:
        findings.append(
            _finding(
                findings,
                event_id=None,
                cte=None,
                severity=summary_finding.get("severity", "medium"),
                status=summary_finding.get("status", "needs_review"),
                finding_type=summary_finding.get("finding_type", "partner_data_quality"),
                message=summary_finding.get("message", ""),
                obligation=fallback_obligation,
                evidence_ids=[],
                confidence=0.85,
                source_citation_override=_regulation_citation("21 CFR 1.1340"),
            )
        )
    return findings


def _regulation_citation(section: str) -> dict[str, Any]:
    return {
        "source_id": "fr-2022-24417-final-rule",
        "citation_anchor": section,
        "section_ref": section,
        "sourceType": "regulation",
    }


def _best_practice_citation() -> dict[str, Any]:
    return {
        "source_id": "traceready-recall-readiness",
        "citation_anchor": "Recall-readiness data-quality check",
        "section_ref": "TraceReady best practice (not a Subpart S requirement)",
        "sourceType": "best_practice",
    }


def _read_inbound_lines(inbound_file: Path) -> tuple[list[dict[str, Any]], str]:
    """Parse one inbound document (EDI 856 / BOL PDF / spreadsheet) into shipment lines."""
    data = inbound_file.read_bytes()
    from bellwether_backend.audit_engine.edi_x12 import edi_856_to_lines, looks_like_x12, parse_x12

    if looks_like_x12(data):
        lines: list[dict[str, Any]] = []
        for transaction in parse_x12(data).transactions:
            if transaction.transaction_set == "856":
                lines.extend(edi_856_to_lines(transaction))
        return lines, f"ASN {inbound_file.name}"
    if data[:5] == b"%PDF-":
        from bellwether_backend.intelligence.bol_extractor import extract_bol_lines

        return extract_bol_lines(data, file_name=inbound_file.name).get("lines", []), f"BOL {inbound_file.name}"
    from bellwether_backend.audit_engine.customer_evidence import _row_facts, read_spreadsheet_evidence

    records = read_spreadsheet_evidence(inbound_file)
    rows = _row_facts(records)
    lines = []
    for position, row in enumerate(sorted(rows.values(), key=lambda r: (r["sheet"], r["row_number"])), start=1):
        facts = {k: [v for v in values if str(v).strip()] for k, values in row["facts"].items() if not k.startswith("source_column:")}
        lines.append({"line_number": position, "facts": {k: v for k, v in facts.items() if v}})
    return lines, f"document {inbound_file.name}"


def _record_gap_message(cte: str | None, missing_fields: list[str], is_tlc_root: bool) -> str:
    label = _cte_label(cte)
    if is_tlc_root:
        other_fields = [field for field in missing_fields if field not in TLC_FIELD_KEYS]
        message = f"{label} is missing its Traceability Lot Code (TLC)"
        if other_fields:
            message += f" and {len(other_fields)} other required field{'s' if len(other_fields) != 1 else ''}"
        return message + "."
    if len(missing_fields) == 1:
        return f"{label} is missing a required field: {_field_label(missing_fields[0])}."
    return f"{label} is missing {len(missing_fields)} required fields."


def _record_sub_issues(missing_fields: list[str], tlc_reasons: list[str]) -> list[str]:
    issues = [f"Missing {_field_label(field)}" for field in missing_fields]
    for reason in tlc_reasons:
        if reason not in issues:
            issues.append(reason)
    return issues


def generate_exception_queue(findings: list[AuditFinding]) -> list[ExceptionQueueItem]:
    queue: list[ExceptionQueueItem] = []
    for finding in findings:
        queue_type = {
            "kde_completeness": "missing_kde",
            "tlc_lineage": "tlc_gap",
            "traceability_plan": "traceability_plan_gap",
            "sortable_export_readiness": "export_blocker",
            "exemption_claim": "exemption_review",
        }.get(finding.finding_type, "review_item")
        queue.append(
            ExceptionQueueItem(
                exception_id=f"phase11-exception-{len(queue) + 1:04d}",
                finding_id=finding.finding_id,
                event_id=finding.event_id,
                queue_type=queue_type,
                priority="high" if finding.severity == "high" else "medium",
                title=f"{finding.finding_type}: {finding.status}",
                details=finding.message,
                evidence_ids=finding.customer_evidence_ids,
                assigned_role="compliance_reviewer",
            )
        )
    return queue


def build_fda_style_export_package(
    *,
    rule_package: dict[str, Any],
    events: dict[str, CustomerEventNode],
    hardened_results: list[MultiSignalCteResult],
    kde_checks: list[KdeCompletenessCheck],
    audit_findings: list[AuditFinding],
    sortable_export_checks: list[SortableExportReadinessCheck],
) -> FdaStyleExportPackage:
    checks_by_event = defaultdict(list)
    for check in kde_checks:
        checks_by_event[(check.event_id, check.cte)].append(check)
    tabs: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for result in hardened_results:
        event = events[result.event_id]
        for cte in result.final_ctes:
            row = {
                "event_id": result.event_id,
                "cte": cte,
                "product_name": event.product_name,
                "traceability_lot_code": event.lot_or_tlc,
                "source_lot_or_tlc": event.source_lot_or_tlc,
                "output_lot_or_tlc": event.output_lot_or_tlc,
                "event_datetime": event.event_datetime,
                "source_evidence_ids": event.evidence_ids,
            }
            for check in checks_by_event.get((result.event_id, cte), []):
                row[check.field_key] = check.observed_values[0] if check.observed_values else None
            tabs[cte].append(row)
    blockers = [
        {
            "check_id": check.check_id,
            "event_id": check.event_id,
            "cte": check.cte,
            "message": f"Sortable export blocked by missing fields: {', '.join(check.missing_fields)}.",
            "missing_fields": check.missing_fields,
        }
        for check in sortable_export_checks
        if check.status != "ready"
    ]
    citations = [
        citation
        for obligation in rule_package["records"]["obligations"]
        for citation in obligation.get("citations", [])[:1]
    ]
    return FdaStyleExportPackage(
        package_id="phase11-fda-style-export-package",
        generated_at=GENERATED_AT,
        rule_package_id=rule_package["package_id"],
        rule_package_version=rule_package["version"],
        status="blocked" if blockers else "ready",
        workbook_file="phase11-fda-style-export-package.xlsx",
        tabs=dict(sorted(tabs.items())),
        blockers=blockers,
        citations=citations,
    )


def _approved_obligations(rule_package: dict[str, Any]) -> dict[str, dict[str, Any]]:
    obligations = {}
    for obligation in rule_package["records"]["obligations"]:
        if obligation.get("metadata", {}).get("review_status") != "approved":
            continue
        obligations[obligation["obligation_id"]] = obligation
    return obligations


def _mapping(event_id: str, cte: str, obligation: dict[str, Any], rule_package: dict[str, Any], index: int) -> EventObligationMapping:
    return EventObligationMapping(
        mapping_id=f"phase11-map-{index:04d}",
        event_id=event_id,
        cte=cte,
        approved_obligation_id=obligation["obligation_id"],
        obligation_action=obligation.get("action", ""),
        required_output=obligation.get("required_output"),
        citation=obligation.get("citations", [{}])[0],
        rule_package_id=rule_package["package_id"],
        rule_package_version=rule_package["version"],
    )


def _promote_kde_value_rows(
    evidence_by_id: dict[str, CustomerEvidenceRecord],
) -> dict[str, dict[str, list[tuple[str, str]]]]:
    """Reconstruct the 07_KDE_Values sheet (one logical KDE per row, declaring its own
    field_key + value) into event_id -> field_key -> [(value, evidence_id)]. This lets a
    contract's satisfied_by reference KDE-sheet fields the customer actually provided."""
    rows: "OrderedDict[Any, dict[str, tuple[str, str]]]" = OrderedDict()
    for record in evidence_by_id.values():
        if record.sheet_name != "07_KDE_Values":
            continue
        rows.setdefault(record.row_number, {})[record.field_key] = (record.normalized_value, record.evidence_id)
    promoted: dict[str, dict[str, list[tuple[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows.values():
        event_id = (row.get("event_id") or ("", ""))[0].strip()
        field_key = (row.get("kde_field_key") or ("", ""))[0].strip()
        value, evidence_id = row.get("kde_value") or ("", "")
        if event_id and field_key and value:
            promoted[event_id][field_key].append((value, evidence_id))
    return promoted


def _merge_promoted_kde_facts(
    facts: dict[str, list[str]],
    promoted: dict[str, dict[str, list[tuple[str, str]]]],
    event_node_id: str,
) -> None:
    """Merge promoted KDE-sheet facts into an event's facts. KDE rows key by the raw event id
    (e.g. SHIP-1); event nodes are scoped by product (SHIP-1:PROD-1), so match on the prefix."""
    raw_event_id = event_node_id.split(":", 1)[0]
    for candidate_event_id, field_values in promoted.items():
        if candidate_event_id != raw_event_id and not event_node_id.startswith(f"{candidate_event_id}:"):
            continue
        for field_key, pairs in field_values.items():
            for value, evidence_id in pairs:
                facts[field_key].append(value)
                if evidence_id:
                    facts[f"evidence:{field_key}"].append(evidence_id)


def _event_facts(event: CustomerEventNode, evidence_by_id: dict[str, CustomerEvidenceRecord]) -> dict[str, list[str]]:
    facts: dict[str, list[str]] = defaultdict(list)
    direct = {
        "event_datetime": event.event_datetime,
        "product_name": event.product_name,
        "traceability_lot_code": event.lot_or_tlc,
        "source_lot_or_tlc": event.source_lot_or_tlc,
        "output_lot_or_tlc": event.output_lot_or_tlc,
    }
    for key, value in direct.items():
        if value:
            facts[key].append(value)
            facts[f"evidence:{key}"].extend(event.evidence_ids)
    for evidence_id in event.evidence_ids:
        record = evidence_by_id.get(evidence_id)
        if not record:
            continue
        facts[record.field_key].append(record.normalized_value)
        facts[f"evidence:{record.field_key}"].append(record.evidence_id)
    return facts


def _find_obligation_by_cte(approved_obligations: dict[str, dict[str, Any]], cte: str) -> dict[str, Any]:
    for obligation in approved_obligations.values():
        if cte in (obligation.get("applies_to_ctes") or []):
            return obligation
    raise ValueError(f"missing approved obligation for CTE {cte}")


def _is_specific_cte_obligation(obligation: dict[str, Any], cte: str) -> bool:
    section_ref = CTE_SECTION_REFS.get(cte)
    if not section_ref:
        return False
    citations = obligation.get("citations") or []
    return any(
        isinstance(citation, dict) and str(citation.get("section_ref") or "").startswith(section_ref)
        for citation in citations
    )


def _find_obligation_containing(approved_obligations: dict[str, dict[str, Any]], token: str) -> dict[str, Any]:
    for obligation in approved_obligations.values():
        if token in obligation["obligation_id"]:
            return obligation
    raise ValueError(f"missing approved obligation containing {token}")


def _finding(
    findings: list[AuditFinding],
    *,
    event_id: str | None,
    cte: str | None,
    severity: str,
    status: str,
    finding_type: str,
    message: str,
    obligation: dict[str, Any],
    evidence_ids: list[str],
    confidence: float,
    sub_issues: list[str] | None = None,
    affected_fields: list[str] | None = None,
    source_citation_override: dict[str, Any] | None = None,
    requirement_source: str = "fda_rule",
) -> AuditFinding:
    return AuditFinding(
        finding_id=f"phase11-finding-{len(findings) + 1:04d}",
        event_id=event_id,
        cte=cte,
        severity=severity,
        status=status,
        finding_type=finding_type,
        message=message,
        approved_obligation_id=obligation["obligation_id"],
        source_citation=source_citation_override if source_citation_override is not None else obligation.get("citations", [{}])[0],
        customer_evidence_ids=evidence_ids,
        confidence=confidence,
        reviewer_status="needs_review" if status in {"gap", "needs_review"} else "system_pass",
        sub_issues=sub_issues or [],
        affected_fields=affected_fields or [],
        requirement_source=requirement_source,
    )


def _summary(
    *,
    rule_package: dict[str, Any],
    obligation_mappings: list[EventObligationMapping],
    kde_checks: list[KdeCompletenessCheck],
    tlc_checks: list[TlcLineageCheck],
    traceability_plan_checks: list[TraceabilityPlanCheck],
    scope_exemption_checks: list[ScopeExemptionCheck],
    records_readiness_checks: list[RecordsReadinessCheck],
    sortable_export_checks: list[SortableExportReadinessCheck],
    audit_findings: list[AuditFinding],
    exception_queue: list[ExceptionQueueItem],
    export_package: FdaStyleExportPackage,
) -> dict[str, Any]:
    finding_status_counts = Counter(finding.status for finding in audit_findings)
    return {
        "phase": 11,
        "generatedAt": GENERATED_AT,
        "rulePackageId": rule_package["package_id"],
        "rulePackageVersion": rule_package["version"],
        "approvedRuleOnly": True,
        "obligationMappings": len(obligation_mappings),
        "kdeChecks": len(kde_checks),
        "kdeStatusCounts": dict(sorted(Counter(check.status for check in kde_checks).items())),
        "tlcChecks": len(tlc_checks),
        "tlcStatusCounts": dict(sorted(Counter(check.status for check in tlc_checks).items())),
        "traceabilityPlanChecks": len(traceability_plan_checks),
        "scopeExemptionChecks": len(scope_exemption_checks),
        "recordsReadinessChecks": len(records_readiness_checks),
        "sortableExportChecks": len(sortable_export_checks),
        "auditFindings": len(audit_findings),
        "findingStatusCounts": dict(sorted(finding_status_counts.items())),
        "exceptionQueueItems": len(exception_queue),
        "exportPackageStatus": export_package.status,
        "acceptanceCoverage": {
            "RI-100_map_customer_ctes_to_approved_obligations": True,
            "RI-101_kde_completeness_checks": True,
            "RI-102_tlc_lineage_checks": True,
            "RI-103_traceability_plan_checks": True,
            "RI-104_exemption_scope_uncertainty_checks": True,
            "RI-105_records_fda_request_readiness": True,
            "RI-106_sortable_export_readiness": True,
            "RI-107_audit_findings": True,
            "RI-108_exception_queue": True,
            "RI-109_fda_style_export_package": True,
        },
    }


def _write_export_workbook(path: Path, export_package: FdaStyleExportPackage, findings: list[AuditFinding]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception:
        return
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["package_id", export_package.package_id])
    summary.append(["status", export_package.status])
    summary.append(["rule_package_id", export_package.rule_package_id])
    summary.append(["rule_package_version", export_package.rule_package_version])
    summary.append(["blockers", len(export_package.blockers)])
    for tab_name, rows in export_package.tabs.items():
        sheet = workbook.create_sheet(tab_name[:31] or "Events")
        headers = sorted({key for row in rows for key in row.keys()})
        sheet.append(headers)
        for row in rows:
            sheet.append([json.dumps(row.get(header)) if isinstance(row.get(header), (list, dict)) else row.get(header) for header in headers])
    findings_sheet = workbook.create_sheet("Findings")
    findings_sheet.append(["finding_id", "event_id", "cte", "severity", "status", "type", "message", "approved_obligation_id"])
    for finding in findings:
        findings_sheet.append([
            finding.finding_id,
            finding.event_id,
            finding.cte,
            finding.severity,
            finding.status,
            finding.finding_type,
            finding.message,
            finding.approved_obligation_id,
        ])
    workbook.save(path)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
