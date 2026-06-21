from __future__ import annotations

from io import BytesIO


def extract_xlsx_sheets(xlsx_bytes: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return [{"sheet": "workbook", "text": xlsx_bytes.decode("utf-8", errors="replace")}]

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(" | ".join(values).rstrip())
        text = "\n".join(rows).strip()
        sheets.append({"sheet": sheet.title, "text": text})
    return sheets

